"""
Excel (XLSX) Parser
====================

Reads Excel workbooks — including proper handling of merged cells — and
returns sheets as structured data (list of row-dicts).

**Requires** ``pandas`` and ``openpyxl``.

Usage:
    from IngestionPipeline.Parsers import xlsx_parser

    result = xlsx_parser.parse("data.xlsx")
    # result is a dict mapping sheet names to lists of row-dicts

    # Single sheet
    result = xlsx_parser.parse("data.xlsx", sheet_name="Sheet1")
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook


def _unmerge_and_fill(path: str, sheet: str) -> pd.DataFrame:
    """Read a sheet and forward-fill values from merged cell ranges."""
    df = pd.read_excel(path, sheet_name=sheet)

    try:
        wb = load_workbook(path)
        ws = wb[sheet]
        headers = list(df.columns)

        for merged in ws.merged_cells.ranges:
            if merged.max_row < 2:
                continue
            cell = ws.cell(merged.min_row, merged.min_col)
            val = cell.value
            if val is None:
                continue
            col = headers[merged.min_col - 1] if merged.min_row > 1 else val
            for row in range(merged.min_row, merged.max_row + 1):
                pr = row - 2
                if 0 <= pr < len(df) and col in df.columns and pd.isna(df.iloc[pr][col]):
                    df.at[pr, col] = val
        wb.close()
    except Exception:
        pass  # fall through to unmerged data

    return df


def parse(
    path: str,
    *,
    sheet_name: Optional[str] = None,
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Parse an Excel file and return its content as structured data.

    Args:
        path:        Path to an ``.xlsx`` file.
        sheet_name:  Process only this sheet.  When ``None`` (default)
                     all sheets are processed.

    Returns:
        A dict mapping each sheet name to a list of row-dicts (one dict
        per row, keys are column headers).

    Raises:
        FileNotFoundError: If *path* does not exist.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {path}")

    xf = pd.ExcelFile(str(p))
    sheets = [sheet_name] if sheet_name else xf.sheet_names
    result: Dict[str, List[Dict[str, Any]]] = {}

    for sheet in sheets:
        if sheet not in xf.sheet_names:
            continue

        df = _unmerge_and_fill(str(p), sheet)
        df = df.dropna(axis=1, how='all')

        # Forward-fill object columns
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].fillna(method='ffill')

        # Convert NaN/NaT to None for clean JSON serialisation
        result[sheet] = df.where(df.notna(), None).to_dict(orient="records")

    return result
