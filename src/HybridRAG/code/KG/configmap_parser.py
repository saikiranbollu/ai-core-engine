"""
ConfigMap Excel Parser
======================

Parses the ConfigMap Excel (TC4XX_SW_MCAL_ConfigMap_*.xlsx) to extract:
  - Device-to-Configuration mappings (which configs run on which devices)
  - Config composition (which .arxml sub-module files comprise each config)

The ConfigMap is a build manifest telling the test infrastructure which
AUTOSAR config fragments to assemble for each test configuration, per device.

Usage::

    from configmap_parser import parse_configmap_workbook

    data = parse_configmap_workbook("path/to/ConfigMap.xlsx", module="ETH_17_LETH")
    # data = {
    #     "device_configs": [
    #         {"config_name": "Eth_17_Leth_Config001", "config_index": "001",
    #          "devices": ["TC4D9_COM", "TC499N_STD"], "is_vm": False,
    #          "config_files": ["Dem_001", "EcuC_001", ...]},
    #         ...
    #     ],
    #     "device_variants": ["TC447_COM", "TC457_RDR", ...],
    #     "config_details": {"Port_001": "MII Mac Port-3 pin functionality", ...},
    #     "metadata": {"total_configs": 96, "total_devices": 7, ...}
    # }
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger("configmap_parser")

# Pattern to extract config index from names like "Eth_17_Leth_Config001"
_CONFIG_INDEX_RE = re.compile(r"Config(\d+)")

# Family mapping: device variant → chip family
DEVICE_FAMILY_MAP = {
    "TC447_COM": "TC44X",
    "TC457_RDR": "TC45X",
    "TC469_STD": "TC46X",
    "TC489_COM": "TC48X",
    "TC499N_STD": "TC49X",
    "TC4D9_COM": "TC4DX",
    "TC4Z9_COM": "TC4ZX",
}

# Reverse mapping: BVEC family → device variants
FAMILY_TO_VARIANTS = {}
for _dev, _fam in DEVICE_FAMILY_MAP.items():
    FAMILY_TO_VARIANTS.setdefault(_fam, []).append(_dev)
# Also handle BVEC's broader family codes (TC49X covers TC489 too based on naming)
# BVEC uses: TC4DX, TC49X, TC45X, TC48X
# TC48X → TC489_COM; TC49X → TC499N_STD; TC45X → TC457_RDR, TC447_COM; TC4DX → TC4D9_COM, TC4Z9_COM


def parse_configmap_workbook(
    xlsx_path: str | Path,
    module: str,
    *,
    main_sheet: Optional[str] = None,
) -> dict:
    """Parse ConfigMap Excel workbook.

    Args:
        xlsx_path: Path to the ConfigMap Excel file.
        module: Module name (e.g., "ETH_17_LETH").
        main_sheet: Override the main sheet name (auto-detected if None).

    Returns:
        Dict with keys: device_configs, device_variants, config_details, metadata.
    """
    xlsx_path = Path(xlsx_path)
    logger.info("Parsing ConfigMap: %s (module=%s)", xlsx_path.name, module)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)

    # Auto-detect main sheet (first non-Config_Details sheet)
    if main_sheet is None:
        for name in wb.sheetnames:
            if name != "Config_Details":
                main_sheet = name
                break
        if main_sheet is None:
            main_sheet = wb.sheetnames[0]

    logger.info("Using main sheet: '%s'", main_sheet)

    # Parse main sheet
    device_configs = _parse_main_sheet(wb[main_sheet], module)

    # Collect unique device variants
    all_devices = set()
    for entry in device_configs:
        all_devices.update(entry["devices"])
    device_variants = sorted(all_devices)

    # Parse Config_Details if present
    config_details = {}
    if "Config_Details" in wb.sheetnames:
        config_details = _parse_config_details(wb["Config_Details"])

    wb.close()

    metadata = {
        "source_file": xlsx_path.name,
        "module": module,
        "main_sheet": main_sheet,
        "total_configs": len(device_configs),
        "total_devices": len(device_variants),
        "total_config_details": len(config_details),
    }

    logger.info(
        "Parsed %d config entries across %d device variants",
        len(device_configs), len(device_variants),
    )

    return {
        "device_configs": device_configs,
        "device_variants": device_variants,
        "config_details": config_details,
        "metadata": metadata,
    }


def _parse_main_sheet(ws, module: str) -> list[dict]:
    """Parse the main sheet (device-to-config mapping).

    Structure:
        Row 1: Section headers ("Configuration File generation", "EPC modification")
        Row 2: Column headers (Device Name, Configuration name, Path, File/s, ...)
        Row 3+: Data rows (pairs: main row + optional continuation path row)
    """
    configs = []
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    i = 0
    while i < len(rows):
        row = rows[i]

        # Skip empty rows
        if not row or not any(row):
            i += 1
            continue

        # A config entry has device name (col A) and config name (col B)
        device_str = row[0]
        config_name = row[1]

        # If no config name, this might be a continuation row or empty — skip
        if not config_name:
            i += 1
            continue

        config_name = str(config_name).strip()

        # Extract config index
        m = _CONFIG_INDEX_RE.search(config_name)
        config_index = m.group(1) if m else None

        # Parse devices (comma-separated)
        devices = []
        if device_str:
            devices = [d.strip() for d in str(device_str).split(",") if d.strip()]

        # Parse config files (col D, comma-separated)
        config_files = []
        files_str = row[3] if len(row) > 3 else None
        if files_str:
            config_files = [f.strip() for f in str(files_str).split(",") if f.strip()]

        # Path (col C)
        path = str(row[2]).strip() if row[2] else None

        # Additional Info (col J, index 9) — check for "VM"
        additional_info = row[9] if len(row) > 9 else None
        is_vm = str(additional_info).strip().upper() == "VM" if additional_info else False

        # Check for continuation row (next row has only path in col C)
        secondary_path = None
        if i + 1 < len(rows):
            next_row = rows[i + 1]
            if next_row and not next_row[0] and not next_row[1] and next_row[2]:
                secondary_path = str(next_row[2]).strip()
                i += 1  # consume continuation row

        configs.append({
            "config_name": config_name,
            "config_index": config_index,
            "devices": devices,
            "config_files": config_files,
            "path": path,
            "secondary_path": secondary_path,
            "is_vm": is_vm,
            "module": module,
        })

        i += 1

    return configs


def _parse_config_details(ws) -> dict:
    """Parse Config_Details sheet — lookup of config file descriptions.

    Returns:
        Dict mapping file ID → description (e.g., "Port_001" → "MII Mac Port-3 pin functionality").
    """
    details = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            key = str(row[0]).strip()
            desc = str(row[1]).strip()
            if key:
                details[key] = desc
    return details


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    if len(sys.argv) < 2:
        print("Usage: python configmap_parser.py <xlsx_path> [--module MODULE]")
        sys.exit(1)

    path = sys.argv[1]
    mod = "ETH_17_LETH"
    if "--module" in sys.argv:
        idx = sys.argv.index("--module")
        mod = sys.argv[idx + 1]

    data = parse_configmap_workbook(path, module=mod)

    print(f"\n{'='*60}")
    print(f"ConfigMap Summary for {mod}")
    print(f"{'='*60}")
    print(f"  Source: {data['metadata']['source_file']}")
    print(f"  Configs: {data['metadata']['total_configs']}")
    print(f"  Devices: {data['metadata']['total_devices']}")
    print(f"  Config details entries: {data['metadata']['total_config_details']}")
    print(f"\n  Device variants: {data['device_variants']}")

    print(f"\n  First 5 config entries:")
    for entry in data["device_configs"][:5]:
        print(f"    {entry['config_name']}: devices={entry['devices']}, "
              f"files={entry['config_files'][:3]}..., vm={entry['is_vm']}")

    if data["config_details"]:
        print(f"\n  Config details (sample):")
        for k, v in list(data["config_details"].items())[:5]:
            print(f"    {k}: {v}")
