"""
Test Specification (TS) Excel Parsers
======================================

Document-agnostic parsers that extract structured test-case nodes from the
MCAL test specification Excel workbook (TC4xx_SW_MCAL_TS_<Module>.xlsx)
for ingestion into the Neo4j knowledge graph.

Supported Node Types (from ontology):
    - TS_FunctionalTestCase     (sheet "Test cases")
    - TS_ConfigTestCase         (sheet "Configuration tests")
    - TS_StaticInterfaceTestCase (sheet "Static source code IF tests")
    - TS_WCETTestCase           (sheet "WCET analysis")
    - TS_TestSpecDocument       (synthetic document-level node)

Detection is **sheet-name-based** — each supported sheet maps to a node type.
Traceability columns are parsed to extract PRQ, SWA, SWUD, and HAZOP
references for downstream relationship creation.

Uses the existing ``xlsx_parser`` from the IngestionPipeline for merged-cell
handling, with additional column-mapping logic for each sheet layout.

Usage::

    from testspec_parsers import parse_testspec_workbook

    nodes_by_type = parse_testspec_workbook(
        xlsx_path="path/to/TC4xx_SW_MCAL_TS_Adc.xlsx",
        module="ADC",
    )
    # nodes_by_type == {
    #     "TS_FunctionalTestCase": [...],
    #     "TS_ConfigTestCase": [...],
    #     "TS_StaticInterfaceTestCase": [...],
    #     "TS_WCETTestCase": [...],
    #     "TS_TestSpecDocument": [...],
    # }
"""

from __future__ import annotations

import logging
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger("testspec_parsers")

# ---------------------------------------------------------------------------
# Regex patterns  (traceability tag extraction)
# ---------------------------------------------------------------------------

# PRQ reference:  AU3GM-PRQ-xxxxx
_PRQ_REF_RE = re.compile(r"AU3GM-PRQ-\d+", re.IGNORECASE)

# Feature GUID reference
_GUID_RE = re.compile(
    r"\{?\s*([0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-"
    r"[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12})\s*\}?",
)

# [req featureID=<tag> parentID=<ids>][/req]
_REQ_TAG_RE = re.compile(
    r"\[req\s+featureID\s*=\s*(?P<fid>[^\s\]]+)\s+"
    r"parentID\s*=\s*(?P<pids>[^\]]+?)\s*\]\s*\[/req\]",
    re.IGNORECASE | re.DOTALL,
)

# HAZOP reference: HAZOP-<MODULE>-FM-<nnn>
_HAZOP_REF_RE = re.compile(r"HAZOP-[A-Z]+-FM-\d+", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Traceability extraction helpers
# ---------------------------------------------------------------------------

def _extract_traceability(raw_trace: Optional[str]) -> dict:
    """
    Parse a traceability cell and return categorised references.

    Returns a dict with keys:
        raw            – original traceability text
        prq_references – list of AU3GM-PRQ-xxxxx IDs
        swa_references – list of SWA feature GUIDs
        swud_references – list of SWUD feature GUIDs
        hazop_references – list of HAZOP-xxx-FM-nnn IDs
        feature_tags   – list of (featureID, category) tuples for
                         relationship edge properties
    """
    result: Dict[str, Any] = {
        "raw": raw_trace or "",
        "prq_references": [],
        "swa_references": [],
        "swud_references": [],
        "hazop_references": [],
        "feature_tags": [],
    }
    if not raw_trace:
        return result

    text = str(raw_trace)

    # Extract each [req featureID=... parentID=...][/req] block
    for m in _REQ_TAG_RE.finditer(text):
        fid = m.group("fid").strip()
        pids_raw = m.group("pids").strip()

        # Classify by feature ID pattern
        fid_upper = fid.upper()
        if "_PRQ_" in fid_upper:
            category = "PRQ"
            prqs = _PRQ_REF_RE.findall(pids_raw)
            result["prq_references"].extend(prqs)
        elif "_SWA_" in fid_upper:
            category = "SWA"
            guids = _GUID_RE.findall(pids_raw)
            result["swa_references"].extend(guids)
        elif "_SWUD_" in fid_upper:
            category = "SWUD"
            guids = _GUID_RE.findall(pids_raw)
            result["swud_references"].extend(guids)
        elif "_HAZOP_" in fid_upper:
            category = "HAZOP"
            hazops = _HAZOP_REF_RE.findall(pids_raw)
            result["hazop_references"].extend(hazops)
        else:
            category = "OTHER"

        result["feature_tags"].append((fid, category))

    # Deduplicate
    result["prq_references"] = sorted(set(result["prq_references"]))
    result["swa_references"] = sorted(set(result["swa_references"]))
    result["swud_references"] = sorted(set(result["swud_references"]))
    result["hazop_references"] = sorted(set(result["hazop_references"]))

    return result


def _clean(val: Any) -> Optional[str]:
    """Convert a cell value to a cleaned string, or None if empty."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    return s if s else None


def _bool_flag(val: Any) -> bool:
    """Convert a cell value to boolean (True / 'TRUE' / 'x' / 'X')."""
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("TRUE", "X", "YES", "1")


# ---------------------------------------------------------------------------
# Sheet: "Test cases"  →  TS_FunctionalTestCase
# ---------------------------------------------------------------------------

def _parse_functional_tests(
    wb,
    module: str,
    source_document: str,
) -> List[dict]:
    """
    Parse the 'Test cases' sheet.

    Excel layout (row 1 = header, row 2 = sub-header for Test Category):
        Col A  : Test Case ID
        Col B  : Unit Test       (bool)
        Col C  : Integration Test (bool)
        Col D  : Software Test   (bool)
        Col E  : Test Design Techniques
        Col F  : Test functionality / Objective
        Col G  : Configuration Plan
        Col H  : Test Procedure
        Col I  : Expected behaviour
        Col J  : Manual/Automation
        Col K  : Traceability
        Col L  : Status of TestCase
        Col M  : Comments
    """
    sheet_name = "Test cases"
    if sheet_name not in wb.sheetnames:
        logger.info("Sheet '%s' not found – skipping functional tests", sheet_name)
        return []

    ws = wb[sheet_name]
    nodes: List[dict] = []
    # Data starts at row 4 (row 1 = main header, row 2 = sub-header, row 3 = first data)
    # But we observed: row 1 = header, row 2 = sub-category header, row 3+ = data
    # Let's detect by looking for first non-empty Test Case ID
    start_row = 3
    for r in range(3, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and str(val).startswith(module.capitalize()[:3]):
            start_row = r
            break

    for row in range(start_row, ws.max_row + 1):
        tc_id = _clean(ws.cell(row, 1).value)
        if not tc_id:
            continue

        # Test category (columns B, C, D)
        categories = []
        if _bool_flag(ws.cell(row, 2).value):
            categories.append("UnitTest")
        if _bool_flag(ws.cell(row, 3).value):
            categories.append("IntegrationTest")
        if _bool_flag(ws.cell(row, 4).value):
            categories.append("SoftwareTest")

        test_objective = _clean(ws.cell(row, 6).value)
        test_procedure = _clean(ws.cell(row, 8).value)
        expected_results = _clean(ws.cell(row, 9).value)
        traceability_raw = _clean(ws.cell(row, 11).value)

        # Parse traceability
        trace = _extract_traceability(traceability_raw)

        node = {
            "test_case_id": tc_id,
            "test_category": categories if categories else ["UnitTest"],
            "test_design_techniques": _clean(ws.cell(row, 5).value),
            "test_objective": test_objective or "",
            "configuration_plan": _clean(ws.cell(row, 7).value),
            "test_procedure": test_procedure or "",
            "expected_results": expected_results or "",
            "automation_method": _clean(ws.cell(row, 10).value),
            "traceability_tags": traceability_raw,
            "prq_references": trace["prq_references"],
            "swa_references": trace["swa_references"],
            "swud_references": trace["swud_references"],
            "hazop_references": trace["hazop_references"],
            "status": _clean(ws.cell(row, 12).value),
            "comments": _clean(ws.cell(row, 13).value),
            "module": module.upper(),
            "source_document": source_document,
        }
        nodes.append(node)

    logger.info("  Parsed %d TS_FunctionalTestCase nodes from '%s'",
                len(nodes), sheet_name)
    return nodes


# ---------------------------------------------------------------------------
# Sheet: "Configuration tests"  →  TS_ConfigTestCase
# ---------------------------------------------------------------------------

def _parse_config_tests(
    wb,
    module: str,
    source_document: str,
) -> List[dict]:
    """
    Parse the 'Configuration tests' sheet.

    Excel layout (row 1 = header, row 2 = sub-header):
        Col A  : Test Case ID
        Col B  : Unit Test       (bool)
        Col C  : Integration Test (bool)
        Col D  : Configuration Parameter/Container
        Col E  : Element Type
        Col F-H: Multiplicity (Min, Max, Default)
        Col I-M: Value (Min, Max, Default, Range, Unit)
        Col N  : Is Editable
        Col O  : Parameter Dependency On
        Col P  : Test Procedure
        Col Q  : Expected Results
        Col R  : Manual/Automation
        Col S  : Traceability
        Col T  : Status
        Col U  : Comments
    """
    sheet_name = "Configuration tests"
    if sheet_name not in wb.sheetnames:
        logger.info("Sheet '%s' not found – skipping config tests", sheet_name)
        return []

    ws = wb[sheet_name]
    nodes: List[dict] = []

    # Data starts after header rows (typically row 3)
    start_row = 2
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and str(val).startswith(module.capitalize()[:3]):
            start_row = r
            break

    for row in range(start_row, ws.max_row + 1):
        tc_id = _clean(ws.cell(row, 1).value)
        if not tc_id:
            continue

        # Test category
        categories = []
        if _bool_flag(ws.cell(row, 2).value):
            categories.append("UnitTest")
        if _bool_flag(ws.cell(row, 3).value):
            categories.append("IntegrationTest")

        config_path = _clean(ws.cell(row, 4).value)
        element_type = _clean(ws.cell(row, 5).value)

        # Multiplicity
        mult_min = _clean(ws.cell(row, 6).value)
        mult_max = _clean(ws.cell(row, 7).value)

        # Value range info
        val_min = _clean(ws.cell(row, 9).value)
        val_max = _clean(ws.cell(row, 10).value)
        val_default = _clean(ws.cell(row, 11).value)
        val_range = _clean(ws.cell(row, 12).value)
        # Build combined value_range string
        value_parts = []
        if val_min:
            value_parts.append(f"min={val_min}")
        if val_max:
            value_parts.append(f"max={val_max}")
        if val_default:
            value_parts.append(f"default={val_default}")
        if val_range:
            value_parts.append(f"range={val_range}")
        value_range = "; ".join(value_parts) if value_parts else None

        test_procedure = _clean(ws.cell(row, 16).value)
        expected_results = _clean(ws.cell(row, 17).value)
        traceability_raw = _clean(ws.cell(row, 19).value)

        # Parse traceability
        trace = _extract_traceability(traceability_raw)

        node = {
            "test_case_id": tc_id,
            "test_category": categories if categories else ["UnitTest"],
            "config_path": config_path or "",
            "element_type": element_type,
            "multiplicity_min": mult_min,
            "multiplicity_max": mult_max,
            "value_range": value_range,
            "is_editable": _clean(ws.cell(row, 14).value),
            "parameter_dependency": _clean(ws.cell(row, 15).value),
            "test_procedure": test_procedure or "",
            "expected_results": expected_results or "",
            "automation_method": _clean(ws.cell(row, 18).value),
            "traceability_tags": traceability_raw,
            "prq_references": trace["prq_references"],
            "swa_references": trace["swa_references"],
            "swud_references": trace["swud_references"],
            "hazop_references": trace["hazop_references"],
            "status": _clean(ws.cell(row, 20).value),
            "comments": _clean(ws.cell(row, 21).value),
            "module": module.upper(),
            "source_document": source_document,
        }
        nodes.append(node)

    logger.info("  Parsed %d TS_ConfigTestCase nodes from '%s'",
                len(nodes), sheet_name)
    return nodes


# ---------------------------------------------------------------------------
# Sheet: "Static source code IF tests"  →  TS_StaticInterfaceTestCase
# ---------------------------------------------------------------------------

def _parse_static_tests(
    wb,
    module: str,
    source_document: str,
) -> List[dict]:
    """
    Parse the 'Static source code IF tests' sheet.

    Excel layout (row 1 = header, row 2 = sub-header):
        Col A  : (numbering / empty)
        Col B  : Test Case ID
        Col C  : Unit Test       (bool)
        Col D  : Integration Test (bool)
        Col E  : Test functionality (contains objective + procedure + expectation)
        Col F  : Manual / Automation
        Col G  : Traceability
        Col H  : Status
        Col I  : Comments
    """
    sheet_name = "Static source code IF tests"
    if sheet_name not in wb.sheetnames:
        logger.info("Sheet '%s' not found – skipping static tests", sheet_name)
        return []

    ws = wb[sheet_name]
    nodes: List[dict] = []

    # Data starts after header rows
    start_row = 3
    for r in range(3, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if val and str(val).startswith(module.capitalize()[:3]):
            start_row = r
            break

    for row in range(start_row, ws.max_row + 1):
        tc_id = _clean(ws.cell(row, 2).value)
        if not tc_id:
            continue

        # Test category
        categories = []
        if _bool_flag(ws.cell(row, 3).value):
            categories.append("UnitTest")
        if _bool_flag(ws.cell(row, 4).value):
            categories.append("IntegrationTest")

        # The functionality column contains objective, procedure, and
        # expectation typically as a combined block of text
        functionality = _clean(ws.cell(row, 5).value) or ""

        # Try to split into objective / procedure / expected
        test_objective = functionality
        test_procedure = ""
        expected_results = ""

        # Common pattern: "Objective:-\n...\nProcedure:-\n...\nExpectation:-\n..."
        obj_match = re.search(
            r"(?:Objective|Test\s+functionality)[\s:\-]*\n(.*?)(?=\nProcedure|\nExpect|\Z)",
            functionality, re.IGNORECASE | re.DOTALL
        )
        proc_match = re.search(
            r"Procedure[\s:\-]*\n(.*?)(?=\nExpect|\Z)",
            functionality, re.IGNORECASE | re.DOTALL
        )
        exp_match = re.search(
            r"Expectation[\s:\-]*\n(.*)",
            functionality, re.IGNORECASE | re.DOTALL
        )

        if obj_match:
            test_objective = obj_match.group(1).strip()
        if proc_match:
            test_procedure = proc_match.group(1).strip()
        if exp_match:
            expected_results = exp_match.group(1).strip()

        traceability_raw = _clean(ws.cell(row, 7).value)
        trace = _extract_traceability(traceability_raw)

        node = {
            "test_case_id": tc_id,
            "test_category": categories if categories else ["UnitTest"],
            "test_objective": test_objective,
            "test_procedure": test_procedure,
            "expected_results": expected_results,
            "automation_method": _clean(ws.cell(row, 6).value),
            "traceability_tags": traceability_raw,
            "prq_references": trace["prq_references"],
            "swa_references": trace["swa_references"],
            "swud_references": trace["swud_references"],
            "hazop_references": trace["hazop_references"],
            "status": _clean(ws.cell(row, 8).value),
            "comments": _clean(ws.cell(row, 9).value),
            "module": module.upper(),
            "source_document": source_document,
        }
        nodes.append(node)

    logger.info("  Parsed %d TS_StaticInterfaceTestCase nodes from '%s'",
                len(nodes), sheet_name)
    return nodes


# ---------------------------------------------------------------------------
# Sheet: "WCET analysis"  →  TS_WCETTestCase
# ---------------------------------------------------------------------------

def _parse_wcet_tests(
    wb,
    module: str,
    source_document: str,
) -> List[dict]:
    """
    Parse the 'WCET analysis' sheet.

    Excel layout (row 1 = header):
        Col A  : SL.No
        Col B  : API
        Col C  : Configuration Description
        Col D  : Scenario
        Col E  : Test Case ID
        Col F  : Configuration (file name)
        Col G  : Data Points
        Col H  : Applicable Autosar
    """
    sheet_name = "WCET analysis"
    if sheet_name not in wb.sheetnames:
        logger.info("Sheet '%s' not found – skipping WCET tests", sheet_name)
        return []

    ws = wb[sheet_name]
    nodes: List[dict] = []

    # Track previous API name for merged-cell forward-fill
    prev_api = None
    prev_config_desc = None
    prev_scenario = None

    for row in range(2, ws.max_row + 1):
        tc_id = _clean(ws.cell(row, 5).value)
        if not tc_id:
            continue

        api_name = _clean(ws.cell(row, 2).value) or prev_api
        config_desc = _clean(ws.cell(row, 3).value) or prev_config_desc
        scenario = _clean(ws.cell(row, 4).value) or prev_scenario
        config_name = _clean(ws.cell(row, 6).value)
        data_point = _clean(ws.cell(row, 7).value)
        autosar_ver = _clean(ws.cell(row, 8).value)

        # Update forward-fill state
        if _clean(ws.cell(row, 2).value):
            prev_api = api_name
        if _clean(ws.cell(row, 3).value):
            prev_config_desc = config_desc
        if _clean(ws.cell(row, 4).value):
            prev_scenario = scenario

        # Build a unique composite ID for WCET rows (same test_case_id
        # can appear for ET and WCET data points)
        composite_id = f"{tc_id}_{data_point.replace(' ', '_')}" if data_point else tc_id

        node = {
            "test_case_id": composite_id,
            "api_name": api_name or "",
            "configuration_description": config_desc or "",
            "scenario": scenario or "",
            "config_name": config_name,
            "data_point": data_point,
            "autosar_version": autosar_ver,
            "module": module.upper(),
            "source_document": source_document,
        }
        nodes.append(node)

    logger.info("  Parsed %d TS_WCETTestCase nodes from '%s'",
                len(nodes), sheet_name)
    return nodes


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_testspec_workbook(
    xlsx_path: str | Path,
    module: str,
) -> Dict[str, List[dict]]:
    """
    Parse a MCAL test specification Excel workbook and return structured
    nodes grouped by type, ready for ingestion into Neo4j.

    Args:
        xlsx_path: Path to the test specification .xlsx file.
        module:    MCAL module name (e.g. 'ADC', 'SPI', 'CAN').

    Returns:
        A dict mapping node type labels to lists of node property dicts:
            {
                "TS_FunctionalTestCase": [...],
                "TS_ConfigTestCase": [...],
                "TS_StaticInterfaceTestCase": [...],
                "TS_WCETTestCase": [...],
                "TS_TestSpecDocument": [...],
            }

    Raises:
        FileNotFoundError: If *xlsx_path* does not exist.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Test spec file not found: {xlsx_path}")

    module = module.upper()
    source_doc = path.stem  # e.g. "TC4xx_SW_MCAL_TS_Adc"

    logger.info("Parsing test spec workbook: %s (module: %s)", path.name, module)

    wb = load_workbook(str(path), data_only=True, read_only=False)

    try:
        # Parse each sheet
        functional = _parse_functional_tests(wb, module, source_doc)
        config = _parse_config_tests(wb, module, source_doc)
        static = _parse_static_tests(wb, module, source_doc)
        wcet = _parse_wcet_tests(wb, module, source_doc)

        result: Dict[str, List[dict]] = {}

        if functional:
            result["TS_FunctionalTestCase"] = functional
        if config:
            result["TS_ConfigTestCase"] = config
        if static:
            result["TS_StaticInterfaceTestCase"] = static
        if wcet:
            result["TS_WCETTestCase"] = wcet

        # Create the document node
        doc_node = {
            "document_name": source_doc,
            "description": (
                f"MCAL Test Specification for {module} module – "
                f"extracted from {path.name}"
            ),
            "total_functional_tests": len(functional),
            "total_config_tests": len(config),
            "total_static_tests": len(static),
            "total_wcet_tests": len(wcet),
            "module": module,
            "source_document": path.name,
        }
        result["TS_TestSpecDocument"] = [doc_node]

        # Summary
        total = sum(len(v) for v in result.values())
        logger.info(
            "Parsed %d total test spec nodes: "
            "Functional=%d, Config=%d, Static=%d, WCET=%d, Doc=1",
            total, len(functional), len(config), len(static), len(wcet),
        )

        return result

    finally:
        wb.close()


# ---------------------------------------------------------------------------
# CLI quick-test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import json as _json

    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(name)s - %(message)s")

    if len(sys.argv) < 2:
        print("Usage: python testspec_parsers.py <path_to_xlsx> [module]")
        sys.exit(1)

    xlsx = sys.argv[1]
    mod = sys.argv[2] if len(sys.argv) > 2 else "ADC"

    parsed = parse_testspec_workbook(xlsx, mod)

    for ntype, items in sorted(parsed.items()):
        print(f"\n{'=' * 60}")
        print(f"  {ntype}: {len(items)} nodes")
        print(f"{'=' * 60}")
        for item in items[:2]:
            # Show first two items as preview
            preview = {}
            for k, v in item.items():
                if isinstance(v, str) and len(v) > 120:
                    preview[k] = v[:120] + "…"
                else:
                    preview[k] = v
            print(f"  {_json.dumps(preview, indent=4, default=str)}")
        if len(items) > 2:
            print(f"  ... and {len(items) - 2} more")
