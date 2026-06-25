"""
Ephemeral Sandbox — Sprint 3 (P1 Feature)
==========================================
Temporary KG (NetworkX) + Vector (in-memory) for document exploration.
Third storage tier: Working Memory → Ephemeral Sandbox → Semantic Memory.

Self-contained: no dependency on OntologyLoader or sentence-transformers.
Falls back to hash-based embeddings for dev; real embedder pluggable.

Components:
  EphemeralGraph   — NetworkX in-memory knowledge graph with keyword index
  EphemeralVectors — In-memory vector store with fallback embedder
  SandboxManager   — Lifecycle tied to sessions (auto-cleanup on session_end)
  SandboxIngester  — File parser dispatch → graph + vector population
  SandboxQuerier   — Unified query interface for GraphRAG integration
"""
from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import threading
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Protocol

import yaml as _yaml

logger = logging.getLogger(__name__)

_ILLD_MODULE_ALIAS_CACHE: Optional[Dict[str, str]] = None

# Config-driven default alpha (MEG_SW-308)
try:
    _cfg_path = Path(__file__).resolve().parents[2] / "HybridRAG" / "config" / "storage_config.yaml"
    with open(_cfg_path, "r", encoding="utf-8") as _fh:
        _DEFAULT_SEARCH_ALPHA: float = float(
            _yaml.safe_load(_fh).get("hybrid_search", {}).get("default_alpha", 0.6)
        )
except Exception:
    _DEFAULT_SEARCH_ALPHA = 0.6


# ═════════════════════════════════════════════════════════════════════════
#  Data Classes
# ═════════════════════════════════════════════════════════════════════════

@dataclass
class Chunk:
    """A text chunk ready for embedding."""
    text: str
    chunk_id: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)
    source_file: str = ""
    section: str = ""

    def __post_init__(self):
        if not self.chunk_id:
            self.chunk_id = f"chunk_{uuid.uuid4().hex[:12]}"


@dataclass
class SearchResult:
    """Unified search result from graph or vector search."""
    node_id: str
    content: str
    score: float
    node_type: str = ""
    origin: str = "ephemeral"  # eph_graph | eph_vector
    metadata: Dict[str, Any] = field(default_factory=dict)


# ═════════════════════════════════════════════════════════════════════════
#  Fallback Embedder (hash-based, for dev without sentence-transformers)
# ═════════════════════════════════════════════════════════════════════════

class _FallbackEmbedder:
    """Deterministic hash-based embedder. NOT for production."""
    def __init__(self, dim: int = 384):
        self._dim = dim

    def embed(self, texts: List[str]) -> List[List[float]]:
        results = []
        for text in texts:
            # H18 fix: Expand hash to fill full dim (384) by repeated hashing
            vec = []
            seed = text.encode()
            while len(vec) < self._dim:
                h = hashlib.sha384(seed).digest()
                vec.extend(float(b) / 255.0 for b in h)
                seed = h  # chain hashes
            vec = vec[:self._dim]
            norm = sum(v * v for v in vec) ** 0.5
            vec = [v / norm for v in vec] if norm > 0 else vec
            results.append(vec)
        return results


class _SentenceTransformerEmbedder:
    """Production embedder using the shared SentenceTransformer singleton.

    Falls back to _FallbackEmbedder if sentence-transformers is not installed.
    """

    def __init__(self):
        self._model = None
        self._fallback = None
        self._initialized = False

    def _init(self):
        if self._initialized:
            return
        self._initialized = True
        try:
            from src.Configuration.embedding_singleton import get_shared_model
            self._model = get_shared_model()
        except Exception:
            pass
        if self._model is None:
            self._fallback = _FallbackEmbedder()
            logger.warning("[Sandbox] sentence-transformers unavailable — using hash fallback")

    def embed(self, texts: List[str]) -> List[List[float]]:
        self._init()
        if self._model is None:
            return self._fallback.embed(texts)
        import numpy as np
        embeddings = self._model.encode(texts, normalize_embeddings=True)
        if isinstance(embeddings, np.ndarray):
            return embeddings.tolist()
        return [e.tolist() if hasattr(e, 'tolist') else list(e) for e in embeddings]


# ═════════════════════════════════════════════════════════════════════════
#  EphemeralGraph — NetworkX in-memory knowledge graph
# ═════════════════════════════════════════════════════════════════════════

class EphemeralGraph:
    """In-memory knowledge graph using NetworkX with keyword index."""

    def __init__(self):
        try:
            import networkx as nx
            self._nx = nx
        except ImportError:
            raise ImportError("networkx required: pip install networkx")
        self._graph = nx.DiGraph()
        self._keyword_index: Dict[str, set] = {}

    def add_node(self, node_type: str, node_id: str, props: Optional[Dict] = None):
        props = props or {}
        props["_node_type"] = node_type
        props["_node_id"] = node_id
        self._graph.add_node(node_id, **props)
        self._index_node(node_id, node_type, props)

    def add_relationship(self, source: str, target: str, rel_type: str, props: Optional[Dict] = None):
        for nid in (source, target):
            if nid not in self._graph:
                self._graph.add_node(nid, _node_type="Unknown", _node_id=nid)
        p = props or {}
        p["_rel_type"] = rel_type
        self._graph.add_edge(source, target, **p)

    def add_nodes_bulk(self, nodes: List[Dict]) -> int:
        count = 0
        for n in nodes:
            nt, nid = n.get("node_type", "Unknown"), n.get("node_id", "")
            if nid:
                props = {k: v for k, v in n.items() if k not in ("node_type", "node_id")}
                self.add_node(nt, nid, props)
                count += 1
        return count

    def add_relationships_bulk(self, rels: List[Dict]) -> int:
        count = 0
        for r in rels:
            src, tgt, rt = r.get("source", ""), r.get("target", ""), r.get("rel_type", "RELATED_TO")
            if src and tgt:
                props = {k: v for k, v in r.items() if k not in ("source", "target", "rel_type")}
                self.add_relationship(src, tgt, rt, props)
                count += 1
        return count

    def _index_node(self, node_id: str, node_type: str, props: Dict):
        tokens = set()
        tokens.add(node_id.lower())
        tokens.add(node_type.lower())
        for part in re.split(r'[_\s]+', node_id):
            tokens.add(part.lower())
            for c in re.findall(r'[A-Z]?[a-z]+|[A-Z]+(?=[A-Z]|$)', part):
                tokens.add(c.lower())
        for key in ("function_name", "name", "description", "register_name",
                     "short_name", "requirement_id", "test_id", "module"):
            val = props.get(key)
            if isinstance(val, str):
                for word in re.split(r'[\s_,;]+', val):
                    if len(word) > 2:
                        tokens.add(word.lower())
        for token in tokens:
            if token:
                self._keyword_index.setdefault(token, set()).add(node_id)

    def keyword_search(self, keywords: List[str], node_types: Optional[List[str]] = None,
                       top_k: int = 20) -> List[SearchResult]:
        scores: Dict[str, int] = {}
        for kw in keywords:
            for nid in self._keyword_index.get(kw.lower(), set()):
                scores[nid] = scores.get(nid, 0) + 1
        results = []
        for nid, hit_count in sorted(scores.items(), key=lambda x: -x[1])[:top_k * 2]:
            data = self._graph.nodes.get(nid, {})
            nt = data.get("_node_type", "Unknown")
            if node_types and nt not in node_types:
                continue
            content = " | ".join(f"{k}: {v}" for k, v in data.items() if not k.startswith("_") and v)
            results.append(SearchResult(
                node_id=nid, content=content[:500], score=hit_count / max(len(keywords), 1),
                node_type=nt, origin="eph_graph", metadata={"hit_count": hit_count},
            ))
        return results[:top_k]

    def get_traceability(self, node_ids: List[str]) -> List[Dict]:
        TRACE_RELS = {"SWA_TRACES_TO", "SWUD_REALIZES_SWA", "TS_VERIFIES",
                      "TRACES_TO", "VERIFIES", "REALIZES", "IMPLEMENTS"}
        chain = []
        for nid in node_ids:
            if nid not in self._graph:
                continue
            for _, tgt, edata in self._graph.out_edges(nid, data=True):
                rt = edata.get("_rel_type", "")
                if rt in TRACE_RELS:
                    chain.append({"from": nid, "to": tgt, "relationship": rt})
        return chain

    def stats(self) -> Dict[str, Any]:
        type_counts: Dict[str, int] = {}
        for _, data in self._graph.nodes(data=True):
            nt = data.get("_node_type", "Unknown")
            type_counts[nt] = type_counts.get(nt, 0) + 1
        return {"total_nodes": self._graph.number_of_nodes(),
                "total_edges": self._graph.number_of_edges(),
                "node_types": type_counts, "keyword_index_size": len(self._keyword_index)}

    def clear(self) -> Dict[str, int]:
        stats = {"nodes_removed": self._graph.number_of_nodes(),
                 "edges_removed": self._graph.number_of_edges()}
        self._graph.clear()
        self._keyword_index.clear()
        return stats

    # ── Phase 8 (Plan 2): Enhanced accessors for graph overlay ────────────

    def get_node(self, node_id: str) -> Optional[Dict[str, Any]]:
        """Return properties of a single node, or None if not found."""
        if node_id not in self._graph:
            return None
        return dict(self._graph.nodes[node_id])

    def update_node(self, node_id: str, properties: Dict[str, Any]):
        """Update existing node properties (preserves edges)."""
        if node_id not in self._graph:
            raise KeyError(f"Node {node_id} not found")
        self._graph.nodes[node_id].update(properties)
        nt = properties.get("_node_type", self._graph.nodes[node_id].get("_node_type", "Unknown"))
        self._index_node(node_id, nt, dict(self._graph.nodes[node_id]))

    def get_edge(self, source: str, target: str, rel_type: Optional[str] = None) -> Optional[Dict[str, Any]]:
        """Return edge data between two nodes, or None if not found."""
        data = self._graph.get_edge_data(source, target)
        if data and (rel_type is None or data.get("_rel_type") == rel_type):
            return dict(data)
        return None

    def update_edge(self, source: str, target: str, rel_type: str, properties: Dict[str, Any]):
        """Update an existing edge's properties."""
        if self._graph.has_edge(source, target):
            self._graph.edges[source, target].update(properties)

    def remove_outgoing_edges_by_type(self, node_id: str, rel_types: set) -> int:
        """Remove outgoing edges of specific types from a node.

        Used to clear stale prod edges for relationship types the sandbox
        parser can reliably re-detect (e.g. SRC_CALLS).  Prod edges for
        types the sandbox CANNOT detect (SRC_ACCESSES_SFR, SRC_USES_GLOBAL)
        are preserved as inherited context.

        Returns the number of edges removed.
        """
        if node_id not in self._graph:
            return 0
        to_remove = [
            (node_id, tgt)
            for _, tgt, edata in self._graph.out_edges(node_id, data=True)
            if edata.get("_rel_type") in rel_types
        ]
        for edge in to_remove:
            self._graph.remove_edge(*edge)
        return len(to_remove)

    def remove_node(self, node_id: str) -> bool:
        """Remove a node and all its incident edges from the graph.

        Also purges the node from the keyword index so keyword search stays
        consistent.  Returns True if the node existed, False otherwise.
        """
        if node_id not in self._graph:
            return False
        self._graph.remove_node(node_id)   # NetworkX removes all incident edges too
        for token_set in self._keyword_index.values():
            token_set.discard(node_id)
        return True

    def get_all_nodes(self):
        """Return iterator of (node_id, data_dict) for all nodes."""
        return self._graph.nodes(data=True)

    def get_nodes_by_origin(self, origin: str) -> List[tuple]:
        """Return [(node_id, data)] filtered by _origin metadata."""
        return [(nid, dict(data)) for nid, data in self._graph.nodes(data=True)
                if data.get("_origin") == origin]

    def get_boundary_nodes(self) -> List[Dict[str, Any]]:
        """Return Unknown-type nodes suitable for cross-module resolution.

        Filters out ALL_CAPS names (macros/HW register accesses) that
        won't have corresponding nodes in the production KG.
        """
        boundary = []
        for nid, data in self._graph.nodes(data=True):
            if data.get("_node_type") != "Unknown":
                continue
            # Extract name from canonical ID: "SRC_Function:Dma_ChUpdate:Adc"
            parts = nid.split(":")
            name = parts[1] if len(parts) >= 2 else nid
            # Skip ALL_CAPS names (macros/HW register accesses)
            if re.match(r'^[A-Z][A-Z0-9_]+$', name):
                continue
            boundary.append({"node_id": nid, "name": name})
        return boundary

    def resolve_boundary(self, name_to_resolved: Dict[str, str]) -> Dict[str, int]:
        """Rewire edges from Unknown stubs to resolved production nodes.

        For each name that was resolved against prod, find the Unknown stub
        in the sandbox, redirect all its edges to the resolved prod node,
        and remove the stub.

        Parameters
        ----------
        name_to_resolved : dict
            Maps function name -> canonical ID of the resolved prod node.

        Returns
        -------
        dict with {"boundary_resolved": int}
        """
        resolved_count = 0
        for name, resolved_id in name_to_resolved.items():
            if resolved_id not in self._graph:
                continue
            # Find Unknown stubs matching this name
            stubs = []
            for nid, data in list(self._graph.nodes(data=True)):
                if data.get("_node_type") != "Unknown":
                    continue
                parts = nid.split(":")
                stub_name = parts[1] if len(parts) >= 2 else nid
                if stub_name == name and nid != resolved_id:
                    stubs.append(nid)
            for stub_id in stubs:
                # Redirect incoming edges
                for pred in list(self._graph.predecessors(stub_id)):
                    edata = dict(self._graph.edges[pred, stub_id])
                    if not self._graph.has_edge(pred, resolved_id):
                        self._graph.add_edge(pred, resolved_id, **edata)
                # Redirect outgoing edges
                for succ in list(self._graph.successors(stub_id)):
                    edata = dict(self._graph.edges[stub_id, succ])
                    if not self._graph.has_edge(resolved_id, succ):
                        self._graph.add_edge(resolved_id, succ, **edata)
                # Remove stub from graph and keyword index
                self._graph.remove_node(stub_id)
                for token_set in self._keyword_index.values():
                    token_set.discard(stub_id)
                resolved_count += 1
        return {"boundary_resolved": resolved_count}

    def load_prod_nodes(self, nodes: List[Dict], relationships: List[Dict]):
        """Load production traceability neighbors into ephemeral graph."""
        id_map = {}  # neo4j elementId → canonical node_id

        for node in nodes:
            canonical = self._canonical_id(node["node_type"], node["properties"])
            id_map[node["node_id"]] = canonical

            props = {**node["properties"],
                     "_origin": "production",
                     "_neo4j_id": node["node_id"]}
            self.add_node(node["node_type"], canonical, props)

        for rel in relationships:
            src = id_map.get(rel["source"])
            tgt = id_map.get(rel["target"])
            if src and tgt:
                self.add_relationship(src, tgt, rel["rel_type"],
                                      {**rel.get("properties", {}),
                                       "_origin": "production"})

    @staticmethod
    def _canonical_id(node_type: str, props: Dict) -> str:
        """Consistent identity matching prod KG naming.

        Module is always uppercased to match Neo4j convention
        (prod KG stores module in UPPERCASE, e.g. 'ADC', 'DMA').
        """
        name = (props.get("name") or props.get("function_name")
                or props.get("param_name") or props.get("requirement_id")
                or props.get("test_case_id") or str(props.get("_neo4j_id", "")))
        module = (props.get("module") or "unknown").upper()
        return f"{node_type}:{name}:{module}"

    @property
    def node_count(self): return self._graph.number_of_nodes()
    @property
    def edge_count(self): return self._graph.number_of_edges()


# ═════════════════════════════════════════════════════════════════════════
#  EphemeralVectors — In-memory vector store
# ═════════════════════════════════════════════════════════════════════════

class EphemeralVectors:
    """In-memory vector store. Uses cosine similarity on raw embeddings."""

    def __init__(self, session_id: str, embedder=None, max_chunks: int = 5000):
        self._session_id = session_id
        self._embedder = embedder or _FallbackEmbedder()
        self._max_chunks = max_chunks
        self._chunks: List[Dict] = []  # {id, text, embedding, metadata}

    def add_chunks(self, chunks: List[Chunk]) -> int:
        if not chunks:
            return 0
        if len(self._chunks) + len(chunks) > self._max_chunks:
            raise ValueError(f"Would exceed sandbox limit ({len(self._chunks)}/{self._max_chunks})")
        texts = [c.text for c in chunks]
        embeddings = self._embedder.embed(texts)
        for chunk, emb in zip(chunks, embeddings):
            self._chunks.append({
                "id": chunk.chunk_id, "text": chunk.text, "embedding": emb,
                "metadata": {**chunk.metadata, "source_file": chunk.source_file, "section": chunk.section},
            })
        return len(chunks)

    def search(self, query: str, top_k: int = 10) -> List[SearchResult]:
        if not self._chunks:
            return []
        q_emb = self._embedder.embed([query])[0]
        scored = []
        for chunk in self._chunks:
            sim = self._cosine_sim(q_emb, chunk["embedding"])
            scored.append((sim, chunk))
        scored.sort(key=lambda x: -x[0])
        results = []
        for sim, chunk in scored[:top_k]:
            results.append(SearchResult(
                node_id=chunk["id"], content=chunk["text"][:500], score=max(0, sim),
                origin="eph_vector",
                metadata={**chunk["metadata"], "_session": self._session_id},
            ))
        return results

    @staticmethod
    def _cosine_sim(a: List[float], b: List[float]) -> float:
        dot = sum(x * y for x, y in zip(a, b))
        na = sum(x * x for x in a) ** 0.5
        nb = sum(x * x for x in b) ** 0.5
        return dot / (na * nb) if na > 0 and nb > 0 else 0.0

    def stats(self) -> Dict[str, Any]:
        return {"chunk_count": len(self._chunks), "max_chunks": self._max_chunks,
                "session_id": self._session_id}

    def clear(self) -> Dict[str, int]:
        n = len(self._chunks)
        self._chunks.clear()
        return {"chunks_removed": n}

    def remove_by_metadata(self, **filters):
        """Remove chunks matching all filter criteria."""
        self._chunks = [
            c for c in self._chunks
            if not all(c["metadata"].get(k) == v for k, v in filters.items())
        ]


# ═════════════════════════════════════════════════════════════════════════
#  EphemeralSandbox Container + SandboxManager
# ═════════════════════════════════════════════════════════════════════════

@dataclass
class EphemeralSandbox:
    """Container for one session's ephemeral stores."""
    session_id: str
    graph: EphemeralGraph
    vectors: EphemeralVectors
    files_ingested: List[Dict[str, Any]] = field(default_factory=list)
    active: bool = True

    def status(self) -> Dict[str, Any]:
        return {"session_id": self.session_id, "active": self.active,
                "files": [f.get("filename", "?") for f in self.files_ingested],
                "file_count": len(self.files_ingested),
                "graph_stats": self.graph.stats(), "vector_stats": self.vectors.stats()}

    def clear(self) -> Dict[str, Any]:
        gs = self.graph.clear()
        vs = self.vectors.clear()
        fc = len(self.files_ingested)
        self.files_ingested.clear()
        self.active = False
        return {"graph": gs, "vectors": vs, "files_cleared": fc}


class SandboxManager:
    """Lifecycle manager — one sandbox per session, auto-cleanup."""

    def __init__(self, embedder=None, max_chunks: int = 5000):
        self._embedder = embedder
        self._max_chunks = max_chunks
        self._sandboxes: Dict[str, EphemeralSandbox] = {}
        self._lock = threading.Lock()

    def create_sandbox(self, session_id: str) -> EphemeralSandbox:
        with self._lock:
            if session_id in self._sandboxes and self._sandboxes[session_id].active:
                return self._sandboxes[session_id]
            sb = EphemeralSandbox(
                session_id=session_id,
                graph=EphemeralGraph(),
                vectors=EphemeralVectors(session_id, self._embedder, self._max_chunks),
            )
            self._sandboxes[session_id] = sb
            logger.info("[SandboxManager] Created sandbox for %s", session_id)
            return sb

    def get_sandbox(self, session_id: str) -> Optional[EphemeralSandbox]:
        with self._lock:
            sb = self._sandboxes.get(session_id)
            return sb if sb and sb.active else None

    def destroy_sandbox(self, session_id: str) -> Dict[str, Any]:
        with self._lock:
            sb = self._sandboxes.pop(session_id, None)
            if not sb:
                return {"cleared": False, "reason": "no sandbox found"}
            stats = sb.clear()
            return {"cleared": True, "session_id": session_id, **stats}

    def get_status(self, session_id: str) -> Dict[str, Any]:
        sb = self.get_sandbox(session_id)
        if not sb:
            return {"active": False, "session_id": session_id, "files": [], "file_count": 0}
        return sb.status()

    @property
    def active_count(self) -> int:
        with self._lock:
            return sum(1 for s in self._sandboxes.values() if s.active)


# ═════════════════════════════════════════════════════════════════════════
#  SandboxIngester — Parse files into graph + vectors
# ═════════════════════════════════════════════════════════════════════════

class SandboxIngester:
    """Dispatch file parsing → populate sandbox graph + vectors."""

    SUPPORTED_EXTENSIONS = {".c", ".h", ".txt", ".md", ".rst", ".csv", ".json", ".pdf", ".xlsx", ".puml", ".arxml"}

    def __init__(self, sandbox: EphemeralSandbox, chunk_size: int = 500):
        self._sandbox = sandbox
        self._chunk_size = chunk_size

    def ingest_files(self, file_paths: List[str]) -> Dict[str, Any]:
        stats = {"files_processed": 0, "nodes_created": 0, "edges_created": 0,
                 "chunks_embedded": 0, "errors": []}
        for fp in file_paths:
            try:
                result = self._ingest_one(fp)
                stats["files_processed"] += 1
                stats["nodes_created"] += result.get("nodes", 0)
                stats["edges_created"] += result.get("edges", 0)
                stats["chunks_embedded"] += result.get("chunks", 0)
                self._sandbox.files_ingested.append({"filename": Path(fp).name, "path": fp, **result})
            except Exception as e:
                stats["errors"].append({"file": fp, "error": str(e)})
                logger.error("[SandboxIngester] Failed %s: %s", fp, e)
        return stats

    def ingest_documents(self, documents: List[Dict[str, str]]) -> Dict[str, Any]:
        """Ingest documents provided as text content (no filesystem access needed).

        Parameters
        ----------
        documents : list of dict
            Each dict must have ``filename`` (str) and ``content`` (str).
        """
        stats = {"files_processed": 0, "nodes_created": 0, "edges_created": 0,
                 "chunks_embedded": 0, "errors": []}
        for doc in documents:
            filename = doc.get("filename", "untitled.txt")
            content = doc.get("content", "")
            if not content:
                stats["errors"].append({"file": filename, "error": "Empty content"})
                continue
            try:
                result = self._ingest_content(filename, content)
                stats["files_processed"] += 1
                stats["nodes_created"] += result.get("nodes", 0)
                stats["edges_created"] += result.get("edges", 0)
                stats["chunks_embedded"] += result.get("chunks", 0)
                self._sandbox.files_ingested.append({"filename": filename, "source": "text_upload", **result})
            except Exception as e:
                stats["errors"].append({"file": filename, "error": str(e)})
                logger.error("[SandboxIngester] Failed %s: %s", filename, e)
        return stats

    def _ingest_one(self, file_path: str) -> Dict[str, Any]:
        p = Path(file_path)
        if not p.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        ext = p.suffix.lower()
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported file type: {ext}")

        # Read file
        if ext in (".pdf", ".xlsx"):
            content = f"[Binary file: {p.name} — full parsing requires specialized parser]"
        else:
            content = p.read_text(encoding="utf-8", errors="replace")

        # Create graph nodes
        file_node_id = f"file:{p.name}"
        self._sandbox.graph.add_node("SourceFile", file_node_id, {
            "name": p.name, "path": str(p), "extension": ext, "size_bytes": p.stat().st_size,
        })

        nodes_created = 1
        edges_created = 0

        # Parse C/H files into function nodes
        if ext in (".c", ".h"):
            nodes, edges = self._parse_c_header(content, p.name)
            nodes_created += self._sandbox.graph.add_nodes_bulk(nodes)
            edges_created += self._sandbox.graph.add_relationships_bulk(edges)

        # Chunk content for vectors
        chunks = self._chunk_text(content, p.name)
        chunks_embedded = self._sandbox.vectors.add_chunks(chunks)

        return {"nodes": nodes_created, "edges": edges_created, "chunks": chunks_embedded}

    def _ingest_content(self, filename: str, content: str) -> Dict[str, Any]:
        """Ingest raw text content under a given filename."""
        ext = Path(filename).suffix.lower() if "." in filename else ".txt"

        file_node_id = f"file:{filename}"
        self._sandbox.graph.add_node("SourceFile", file_node_id, {
            "name": filename, "path": "(uploaded text)", "extension": ext,
            "size_bytes": len(content.encode("utf-8")),
        })

        nodes_created = 1
        edges_created = 0

        if ext in (".c", ".h"):
            nodes, edges = self._parse_c_header(content, filename)
            nodes_created += self._sandbox.graph.add_nodes_bulk(nodes)
            edges_created += self._sandbox.graph.add_relationships_bulk(edges)

        chunks = self._chunk_text(content, filename)
        chunks_embedded = self._sandbox.vectors.add_chunks(chunks)

        return {"nodes": nodes_created, "edges": edges_created, "chunks": chunks_embedded}

    def _parse_c_header(self, content: str, filename: str) -> tuple:
        """Extract function declarations from C/H files."""
        nodes, edges = [], []
        # Match function declarations: return_type function_name(params)
        fn_pattern = re.compile(
            r'(?:IFX_EXTERN\s+|IFX_INLINE\s+)?(\w[\w\s\*]+?)\s+(\w+)\s*\(([^)]*)\)\s*[;{]',
            re.MULTILINE
        )
        for m in fn_pattern.finditer(content):
            ret_type, fn_name, params = m.group(1).strip(), m.group(2), m.group(3).strip()
            if fn_name.startswith("_") or len(fn_name) < 3:
                continue
            nodes.append({
                "node_type": "APIFunction", "node_id": fn_name,
                "function_name": fn_name, "return_type": ret_type,
                "parameters": params, "source_file": filename,
            })
            edges.append({"source": fn_name, "target": f"file:{filename}", "rel_type": "DEFINED_IN"})
        return nodes, edges

    def _chunk_text(self, content: str, filename: str) -> List[Chunk]:
        chunks = []
        lines = content.split("\n")
        buf, section = [], ""
        for line in lines:
            if re.match(r'^(#{1,3}|/\*\*|\*\s+@|//\s*={3,})', line):
                if buf:
                    chunks.append(Chunk(
                        text="\n".join(buf), source_file=filename, section=section,
                    ))
                    buf = []
                section = line.strip()[:80]
            buf.append(line)
            if len("\n".join(buf)) > self._chunk_size:
                chunks.append(Chunk(text="\n".join(buf), source_file=filename, section=section))
                buf = []
        if buf:
            chunks.append(Chunk(text="\n".join(buf), source_file=filename, section=section))
        return chunks


# ═════════════════════════════════════════════════════════════════════════
#  SandboxParserDispatcher — Plan 2 Phase 1
#  Routes files to IngestionPipeline parsers without IngestionService
# ═════════════════════════════════════════════════════════════════════════

class SandboxParserDispatcher:
    """Route files by extension to real IngestionPipeline parsers.

    Returns the standardized dict format used by IngestionService._parse_file():
      {"type": str, "functions": [...], "pages": [...], "file": str, ...}
    """

    SUPPORTED_EXTENSIONS = {
        ".c", ".h", ".txt", ".md", ".rst", ".csv", ".json",
        ".pdf", ".xlsx", ".puml", ".arxml",
    }
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

    def __init__(self, include_paths: Optional[List[str]] = None,
                 skip_default_stubs: bool = False,
                 workspace_id: str = "mcal"):
        self._include_paths = include_paths or []
        self._skip_default_stubs = skip_default_stubs
        self._workspace_id = workspace_id

    def parse(self, file_path: Path) -> Dict[str, Any]:
        """Parse a file and return standardized dict."""
        p = Path(file_path) if not isinstance(file_path, Path) else file_path
        ext = p.suffix.lower()
        fname_lower = p.name.lower()

        if ext in (".c", ".h"):
            return self._parse_c(p, ext, fname_lower)
        elif ext == ".pdf":
            return self._parse_pdf(p)
        elif ext == ".xlsx":
            return self._parse_xlsx(p, fname_lower)
        elif ext == ".arxml":
            return self._parse_arxml(p)
        elif ext == ".puml":
            return self._parse_puml(p)
        elif ext == ".rst":
            return self._parse_rst(p)
        elif ext == ".json":
            return self._parse_json(p)
        elif ext in (".txt", ".md", ".csv"):
            return self._parse_text(p, ext)
        else:
            return {"type": "generic", "content": "", "file": str(p)}

    def _parse_c(self, p: Path, ext: str, fname_lower: str) -> Dict:
        # Route iLLD-specific header patterns to dedicated parsers
        if self._workspace_id == "illd" and ext == ".h":
            if "_swa" in fname_lower:
                return self._parse_illd_swa(p)
            if "_regdef" in fname_lower:
                return self._parse_sfr(p)
        try:
            from src.IngestionPipeline.parsers.c_parser import parse as c_parse
            parsed = c_parse(
                str(p),
                include_paths=self._include_paths or None,
                skip_default_stubs=self._skip_default_stubs,
            )
            if isinstance(parsed, dict):
                parsed.setdefault("type", "c_source" if ext == ".c" else "c_header")
                parsed.setdefault("file", str(p))
                # Ensure raw content is available for vector chunking
                if "content" not in parsed:
                    parsed["content"] = p.read_text(encoding="utf-8", errors="replace")
            return parsed
        except ImportError:
            # Fallback to basic regex extraction
            content = p.read_text(encoding="utf-8", errors="replace")
            return {"type": "c_source" if ext == ".c" else "c_header",
                    "functions": [], "content": content, "file": str(p)}

    def _parse_pdf(self, p: Path) -> Dict:
        # Uses pymupdf4llm-based section parser (no GPT, offline-capable).
        # Returns "pdf_md" type → SandboxAdapter creates per-section vector chunks only.
        try:
            from src.IngestionPipeline.parsers.offline_pdf_parser import (
                convert_pdf_to_sections,
            )
            return convert_pdf_to_sections(str(p))
        except ImportError:
            logger.warning("[SandboxParserDispatcher] offline_pdf_parser not available for %s", p.name)
        except Exception as _exc:
            logger.warning("[SandboxParserDispatcher] PDF parser failed for %s: %s", p.name, _exc)

        return {"type": "pdf_md", "sections": [], "raw_md": "", "file": str(p)}

    def _parse_xlsx(self, p: Path, fname_lower: str) -> Dict:
        if "_ts_" in fname_lower or "testspec" in fname_lower:
            try:
                from src.IngestionPipeline.parsers.testspec_parsers import parse_testspec_workbook
                nodes = parse_testspec_workbook(str(p))
                return {"type": "testspec", "nodes": nodes, "file": str(p)}
            except (ImportError, Exception):
                pass

        # iLLD Jama xlsx export detection (workspace_id == "illd")
        if self._workspace_id == "illd":
            result = self._try_parse_illd_jama_xlsx(p)
            if result is not None:
                return result

        try:
            from src.IngestionPipeline.parsers.xlsx_parser import parse as xlsx_parse
            sheets = xlsx_parse(str(p))
            return {"type": "xlsx", "sheets": sheets, "file": str(p)}
        except ImportError:
            return {"type": "xlsx", "sheets": {}, "file": str(p)}

    def _try_parse_illd_jama_xlsx(self, p: Path) -> Optional[Dict]:
        """Detect and parse iLLD Jama requirement export xlsx.

        Format: R3/C1=module, R4=headers (ID,...), R5+=data rows,
        C1=Jama ID (AURC1-REQA-*), C7=Name, C12=Status.
        Sentinel: row starting with "Total Items:".
        """
        try:
            import openpyxl
        except ImportError:
            return None

        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            ws = wb.active
            # Check header row (row 4) — col 1 should be "ID"
            header_val = ws.cell(row=4, column=1).value
            if not header_val or str(header_val).strip().upper() != "ID":
                wb.close()
                return None

            module_name = str(ws.cell(row=3, column=1).value or "").strip()
            requirements = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row or not row[0]:
                    continue
                cell0 = str(row[0]).strip()
                if cell0.startswith("Total Items:"):
                    break
                req_id = cell0
                # C7 = Name (index 6), C12 = Status (index 11)
                name = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                status = str(row[11]).strip() if len(row) > 11 and row[11] else ""
                requirements.append({
                    "requirement_id": req_id,
                    "name": name,
                    "status": status,
                })
            wb.close()
            return {
                "type": "illd_xlsx_req",
                "requirements": requirements,
                "module": module_name,
                "file": str(p),
            }
        except Exception:
            return None

    def _parse_arxml(self, p: Path) -> Dict:
        try:
            from src.IngestionPipeline.parsers.arxml_parser import parse as arxml_parse
            return arxml_parse(str(p))
        except ImportError:
            return {"type": "arxml", "modules": [], "file": str(p)}

    def _parse_puml(self, p: Path) -> Dict:
        try:
            from src.IngestionPipeline.parsers.puml_parser import parse as puml_parse
            return puml_parse(str(p))
        except ImportError:
            return {"type": "puml", "functions": [], "file": str(p)}

    def _parse_rst(self, p: Path) -> Dict:
        try:
            from src.IngestionPipeline.parsers.rst_parser import parse as rst_parse
            sections = rst_parse(str(p))
            return {"type": "rst", "sections": sections, "file": str(p)}
        except ImportError:
            content = p.read_text(encoding="utf-8", errors="replace")
            return {"type": "rst", "content": content, "file": str(p)}

    def _parse_json(self, p: Path) -> Dict:
        import json as _json
        content = p.read_text(encoding="utf-8", errors="replace")
        try:
            data = _json.loads(content)
        except _json.JSONDecodeError:
            data = {}
        return {"type": "json", "data": data, "file": str(p)}

    def _parse_illd_swa(self, p: Path) -> Dict:
        """Route SWA headers to illd_swa_parser (regex-based, no includes)."""
        try:
            from src.IngestionPipeline.parsers.illd_swa_parser import parse as swa_parse
            result = swa_parse(str(p), enrich=False)
            result["type"] = "illd_swa"
            result.setdefault("file", str(p))
            # Attach raw content for vector chunking
            result["content"] = p.read_text(encoding="utf-8", errors="replace")
            return result
        except ImportError:
            content = p.read_text(encoding="utf-8", errors="replace")
            return {"type": "illd_swa", "content": content, "file": str(p),
                    "functions": [], "structs": [], "enums": [],
                    "typedefs": [], "macros": []}

    def _parse_sfr(self, p: Path) -> Dict:
        """Route SFR regdef headers to sfr_parser (regex-based, no includes)."""
        try:
            from src.IngestionPipeline.parsers.sfr_parser import parse as sfr_parse
            result = sfr_parse(str(p))
            result["type"] = "illd_sfr"
            result.setdefault("file", str(p))
            result["content"] = p.read_text(encoding="utf-8", errors="replace")
            return result
        except ImportError:
            content = p.read_text(encoding="utf-8", errors="replace")
            return {"type": "illd_sfr", "content": content, "file": str(p),
                    "registers": {}}

    def _parse_text(self, p: Path, ext: str) -> Dict:
        content = p.read_text(encoding="utf-8", errors="replace")
        return {"type": "text", "content": content, "file": str(p)}


# ═════════════════════════════════════════════════════════════════════════
#  SandboxAdapter — Plan 2 Phase 1 / Phase 5
#  Transforms parser output → EphemeralGraph nodes + EphemeralVectors chunks
#  Implements shadow/override logic for prod node conflicts (Phase 5)
# ═════════════════════════════════════════════════════════════════════════

class SandboxAdapter:
    """Transform parser output into EphemeralGraph nodes/edges + EphemeralVectors chunks.

    When a parsed node conflicts with a production node already in the graph,
    the sandbox version *shadows* it (preserving the original properties).
    """

    def __init__(self, chunk_size: int = 500, workspace_id: str = "mcal"):
        self._chunk_size = chunk_size
        self._workspace_id = workspace_id

    def _extract_node_names_from_parsed(self, parsed: Dict) -> List[str]:
        """Extract node names from parsed data WITHOUT ingesting into sandbox.
        
        Used by sandbox_upload to collect names for traceability pull before
        ingesting (Plan 2 Phase 5: correct load order).
        """
        ptype = parsed.get("type", "")
        node_names = []
        
        if ptype in ("c_source", "c_header"):
            for fn in self._iter_c_functions(parsed):
                name = fn.get("name")
                if name:
                    node_names.append(name)
        elif ptype == "illd_swa":
            for fn in parsed.get("functions", []):
                if isinstance(fn, dict) and fn.get("name"):
                    node_names.append(fn["name"])
            for s in parsed.get("structs", []):
                if isinstance(s, dict) and s.get("name"):
                    node_names.append(s["name"])
            for e in parsed.get("enums", []):
                if isinstance(e, dict) and e.get("name"):
                    node_names.append(e["name"])
            for td in parsed.get("typedefs", []):
                if isinstance(td, dict) and td.get("name"):
                    node_names.append(td["name"])
            for macro in parsed.get("macros", []):
                if isinstance(macro, dict) and macro.get("name"):
                    node_names.append(macro["name"])
        elif ptype == "illd_sfr":
            for reg_name in parsed.get("registers", {}):
                node_names.append(reg_name)
        elif ptype == "illd_xlsx_req":
            for req in parsed.get("requirements", []):
                if isinstance(req, dict):
                    rid = req.get("requirement_id")
                    if rid:
                        node_names.append(rid)
        elif ptype == "json":
            data = parsed.get("data", {})
            reqs = data.get("requirements") if isinstance(data, dict) else None
            if isinstance(reqs, list):
                for req in reqs:
                    if isinstance(req, dict):
                        req_id = req.get("requirement_id") or req.get("id") or req.get("document_key")
                        if req_id:
                            node_names.append(req_id)
        elif ptype == "testspec":
            for tc in parsed.get("nodes", []):
                if isinstance(tc, dict):
                    tc_id = tc.get("test_case_id") or tc.get("id") or tc.get("name")
                    if tc_id:
                        node_names.append(tc_id)
        
        return node_names

    def ingest_parsed(self, sandbox: 'EphemeralSandbox', parsed: Dict, filename: str,
                      module: Optional[str] = None,
                      has_include_paths: bool = False):
        """Convert one parser result dict into sandbox graph nodes + vector chunks."""
        ptype = parsed.get("type", "")
        node_names = []
        illd_semantic_chunks = False  # Flag to skip generic chunking for iLLD

        if ptype == "illd_swa":
            node_names = self._ingest_swa_header(sandbox, parsed, filename, module)
            illd_semantic_chunks = True
        elif ptype == "illd_sfr":
            node_names = self._ingest_sfr(sandbox, parsed, filename, module)
            illd_semantic_chunks = True
        elif ptype == "illd_xlsx_req":
            node_names = self._ingest_xlsx_requirements_illd(sandbox, parsed, filename, module)
            illd_semantic_chunks = True
        elif ptype in ("c_source", "c_header") and self._workspace_id == "illd":
            node_names = self._ingest_c_illd(sandbox, parsed, filename, module)
            illd_semantic_chunks = True
        elif ptype in ("c_source", "c_header"):
            node_names = self._ingest_c(sandbox, parsed, filename, module,
                                        has_include_paths=has_include_paths)
        elif ptype == "json":
            if self._workspace_id == "illd":
                node_names = self._ingest_requirements_illd(sandbox, parsed, filename, module)
            else:
                node_names = self._ingest_json(sandbox, parsed, filename, module)
        elif ptype == "pdf_md":
            node_names = self._ingest_pdf_md(sandbox, parsed, filename, module)
            illd_semantic_chunks = True  # per-section chunks already created
        elif ptype == "xlsx":
            self._ingest_xlsx(sandbox, parsed, filename, module)
        elif ptype == "testspec":
            node_names = self._ingest_testspec(sandbox, parsed, filename, module)
        elif ptype == "arxml":
            self._ingest_arxml(sandbox, parsed, filename, module)

        # Create vector chunks — iLLD uses semantic per-entity chunks (already created
        # inside the _ingest_*_illd methods), so skip the generic text splitter.
        if not illd_semantic_chunks:
            content = self._extract_text_content(parsed)
            if content:
                chunks = self._chunk_text(content, filename)
                sandbox.vectors.remove_by_metadata(source_file=filename)
                sandbox.vectors.add_chunks(chunks)

        # Deletion: remove production nodes from this file that were NOT
        # re-created by this ingest pass (= deleted in the developer's edit).
        # Only run for iLLD types where the ingest completely replaces the
        # file's content.  Vectors are already clean (remove_by_metadata was
        # called at the start of each _ingest_*_illd method).
        deleted_count = 0
        if illd_semantic_chunks and ptype != "pdf_md":
            deleted_count = self._remove_deleted_prod_nodes(sandbox, filename)

        sandbox.files_ingested.append({
            "filename": filename, "source": "parser_dispatch",
            "type": ptype, "node_names_extracted": len(node_names),
            "prod_nodes_deleted": deleted_count,
        })
        return node_names

    def _ingest_c(self, sandbox, parsed, filename, module,
                   has_include_paths: bool = False) -> List[str]:
        """Ingest C/H parsed functions into graph."""
        node_names = []
        mod = (module or "unknown").upper()  # Match prod KG convention (always uppercase)
        # When include_paths were provided, clang detects SFR+globals reliably
        # so we should clear those prod edge types and re-create from sandbox.
        clear_types = (
            self.SANDBOX_DETECTABLE_REL_TYPES_WITH_INCLUDES
            if has_include_paths
            else self.SANDBOX_DETECTABLE_REL_TYPES
        )
        for fn in self._iter_c_functions(parsed):
            name = fn.get("name")
            if not name:
                continue
            node_type = "SRC_Function"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": name, "module": mod})
            # Normalize parameters to match prod format: json.dumps(["type name", ...])
            raw_params = fn.get("parameters", "")
            if isinstance(raw_params, list):
                # Convert [{"name": "x", "type": "int"}, ...] → ["int x", ...]
                param_strs = []
                for p in raw_params:
                    if isinstance(p, dict):
                        param_strs.append(f"{p.get('type', '')} {p.get('name', '')}".strip())
                    else:
                        param_strs.append(str(p))
                params_serialized = json.dumps(param_strs) if param_strs else None
            elif raw_params:
                params_serialized = raw_params if raw_params.startswith("[") else json.dumps([raw_params])
            else:
                params_serialized = None
            props = {
                "name": name,
                "module": mod,
                "return_type": fn.get("return_type", ""),
                "parameters": params_serialized,
                "source_file": filename,
                "function_name": name,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props,
                                       clear_rel_types=clear_types)
            node_names.append(name)

            # Internal call relationships (use SRC_CALLS to match prod naming)
            for call_entry in fn.get("internal_calls", []):
                # Handle both formats: dict {"function": name, ...} / switch-case or plain string
                if isinstance(call_entry, dict):
                    if "function" in call_entry:
                        callee = call_entry["function"]
                    elif "calls" in call_entry:  # switch-case format
                        for sub in call_entry.get("calls", []):
                            callee_name = sub.get("function") if isinstance(sub, dict) else sub
                            if callee_name and callee_name != name:
                                cid = EphemeralGraph._canonical_id("SRC_Function", {"name": callee_name, "module": mod})
                                sandbox.graph.add_relationship(node_id, cid, "SRC_CALLS",
                                                               {"_origin": "sandbox"})
                        continue
                    else:
                        continue
                else:
                    callee = str(call_entry)
                if callee and callee != name:
                    callee_id = EphemeralGraph._canonical_id("SRC_Function", {"name": callee, "module": mod})
                    sandbox.graph.add_relationship(node_id, callee_id, "SRC_CALLS",
                                                   {"_origin": "sandbox"})

            # SFR access relationships (available when include_paths provided)
            for sfr_entry in fn.get("sfr_accesses", []):
                if isinstance(sfr_entry, dict):
                    reg_name = sfr_entry.get("register") or sfr_entry.get("name") or ""
                    if reg_name:
                        sfr_id = EphemeralGraph._canonical_id(
                            "SFR_Register", {"name": reg_name, "module": mod})
                        # Create the SFR_Register node explicitly (not as Unknown)
                        if sandbox.graph.get_node(sfr_id) is None:
                            sandbox.graph.add_node("SFR_Register", sfr_id, {
                                "name": reg_name,
                                "module": mod,
                                "register_name": reg_name,
                                "_origin": "sandbox",
                            })
                        edge_props = {"_origin": "sandbox"}
                        access_type = sfr_entry.get("access_type")
                        if access_type:
                            edge_props["access_type"] = access_type
                        field = sfr_entry.get("field")
                        if field:
                            edge_props["field"] = field
                        sandbox.graph.add_relationship(
                            node_id, sfr_id, "SRC_ACCESSES_SFR", edge_props)

            # Global variable reference relationships (available when include_paths provided)
            for gref in fn.get("global_refs", []):
                if isinstance(gref, dict):
                    gname = gref.get("name") or gref.get("variable") or ""
                    if gname:
                        gid = EphemeralGraph._canonical_id(
                            "SRC_GlobalVariable", {"name": gname, "module": mod})
                        # Create the SRC_GlobalVariable node explicitly (not as Unknown)
                        if sandbox.graph.get_node(gid) is None:
                            sandbox.graph.add_node("SRC_GlobalVariable", gid, {
                                "name": gname,
                                "module": mod,
                                "_origin": "sandbox",
                            })
                        edge_props = {"_origin": "sandbox"}
                        access = gref.get("access_type")
                        if access:
                            edge_props["access_type"] = access
                        sandbox.graph.add_relationship(
                            node_id, gid, "SRC_USES_GLOBAL", edge_props)

        return node_names

    def _ingest_json(self, sandbox, parsed, filename, module) -> List[str]:
        """Ingest JSON (requirements or data)."""
        node_names = []
        mod = module or "unknown"
        data = parsed.get("data", {})
        reqs = data.get("requirements") if isinstance(data, dict) else None
        if isinstance(reqs, list):
            for req in reqs:
                if not isinstance(req, dict):
                    continue
                req_id = req.get("requirement_id") or req.get("id") or req.get("document_key")
                if not req_id:
                    continue
                node_type = "SoftwareRequirement"
                node_id = EphemeralGraph._canonical_id(node_type, {"requirement_id": req_id, "module": mod})
                props = {
                    "requirement_id": req_id,
                    "name": req.get("name") or req_id,
                    "description": req.get("text") or req.get("description") or "",
                    "module": mod,
                    "source_file": filename,
                }
                self._add_node_with_shadow(sandbox, node_type, node_id, props)
                node_names.append(req_id)
        return node_names

    def _ingest_pdf_md(self, sandbox, parsed, filename, module) -> List[str]:
        """Ingest a fast-parsed PDF (type="pdf_md") as in-memory vector chunks only.

        Touches ONLY the in-memory vector store (no graph nodes, no KG edges).
        Creates one semantic chunk per section: heading + content (up to 2000 chars).
        """
        mod = module or "unknown"
        sandbox.vectors.remove_by_metadata(source_file=filename)

        for sec in parsed.get("sections", []):
            heading = sec.get("heading", "")
            if not heading:
                continue
            section_num = sec.get("section_num", "")
            content = sec.get("content", "")
            chunk_text = f"{heading}\n\n{content}".strip()
            sandbox.vectors.add_chunks([Chunk(
                text=chunk_text[:2000],
                source_file=filename,
                metadata={
                    "source_file": filename,
                    "module": mod,
                    "section_num": section_num,
                    "heading": heading,
                },
            )])

        return []

    def _ingest_xlsx(self, sandbox, parsed, filename, module):
        """Ingest XLSX sheets."""
        mod = module or "unknown"
        for sheet_name, rows in (parsed.get("sheets") or {}).items():
            node_type = "Sheet"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": sheet_name, "module": mod})
            props = {"name": sheet_name, "source_file": filename,
                     "module": mod, "row_count": len(rows)}
            self._add_node_with_shadow(sandbox, node_type, node_id, props)

    def _ingest_testspec(self, sandbox, parsed, filename, module) -> List[str]:
        """Ingest test specification nodes."""
        node_names = []
        mod = module or "unknown"
        for tc in parsed.get("nodes", []):
            if not isinstance(tc, dict):
                continue
            tc_id = tc.get("test_case_id") or tc.get("id") or tc.get("name")
            if not tc_id:
                continue
            node_type = "TS_FunctionalTestCase"
            node_id = EphemeralGraph._canonical_id(node_type, {"test_case_id": tc_id, "module": mod})
            props = {**tc, "module": mod, "source_file": filename, "test_case_id": tc_id}
            self._add_node_with_shadow(sandbox, node_type, node_id, props)
            node_names.append(tc_id)
        return node_names

    def _ingest_arxml(self, sandbox, parsed, filename, module):
        """Ingest ARXML modules."""
        mod = module or "unknown"
        for m in parsed.get("modules", []):
            mod_name = m.get("name", "")
            if not mod_name:
                continue
            node_type = "ARXMLModule"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": mod_name, "module": mod})
            props = {"name": mod_name, "source_file": filename, "module": mod}
            self._add_node_with_shadow(sandbox, node_type, node_id, props)

    # ── iLLD-specific Ingestion Methods ────────────────────────────────

    def _ingest_c_illd(self, sandbox, parsed, filename, module) -> List[str]:
        """Ingest C/H parsed functions into iLLD-typed graph nodes.

        Creates Function nodes with CALLS_INTERNALLY edges (matching prod KG).
        """
        node_names = []
        mod = module or "unknown"

        for fn in self._iter_c_functions(parsed):
            name = fn.get("name")
            if not name:
                continue
            node_type = "Function"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": name, "module": mod})
            props = {
                "name": name,
                "id": f"FUNC_{name}",
                "module": mod,
                "return_type": fn.get("return_type", ""),
                "parameters": fn.get("parameters", ""),
                "source_file": filename,
                "source": "Source_Code",
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props,
                                       clear_rel_types=self.ILLD_DETECTABLE_REL_TYPES)
            node_names.append(name)

            # Call graph edges (CALLS_INTERNALLY matching prod)
            for call_entry in fn.get("internal_calls", []):
                if isinstance(call_entry, dict):
                    callee = call_entry.get("function", "")
                else:
                    callee = str(call_entry)
                if callee and callee != name:
                    callee_id = EphemeralGraph._canonical_id(
                        "Function", {"name": callee, "module": mod})
                    sandbox.graph.add_relationship(
                        node_id, callee_id, "CALLS_INTERNALLY",
                        {"_origin": "sandbox"})

        # Semantic per-entity vector chunks for iLLD
        sandbox.vectors.remove_by_metadata(source_file=filename)
        for fn in self._iter_c_functions(parsed):
            name = fn.get("name")
            if not name:
                continue
            text = f"Function: {name}\nReturn: {fn.get('return_type', 'void')}\nParams: {fn.get('parameters', '')}"
            chunk = Chunk(
                text=text,
                source_file=filename,
                metadata={"source_file": filename, "node_type": "Function",
                          "node_id": f"Function:{name}:{mod.upper()}", "module": mod},
            )
            sandbox.vectors.add_chunks([chunk])

        return node_names

    def _ingest_swa_header(self, sandbox, parsed, filename, module) -> List[str]:
        """Ingest SWA parser output → Function, Struct, StructMember, Enum, EnumValue, Typedef nodes.

        Matches production node types, IDs, properties and relationships from illd_kg_builder.py.
        LLM-enriched fields (purpose on Struct/Enum/Typedef, usage_notes, error_handling on
        Function) are stored as empty strings — they require the offline enrichment pipeline.
        Function.purpose uses detailed_description as fallback (same as production fallback).
        """
        _PRIMITIVE_TYPES = {
            "uint8", "uint16", "uint32", "uint64",
            "int8", "int16", "int32", "int64",
            "uint8_t", "uint16_t", "uint32_t", "uint64_t",
            "int8_t", "int16_t", "int32_t", "int64_t",
            "float32", "float64", "boolean", "sint8", "sint16", "sint32",
            "void", "char", "int", "unsigned", "float", "double",
        }

        def _clean_type(raw: str) -> str:
            """Strip pointer/const/volatile qualifiers for type resolution."""
            return raw.replace("const ", "").replace("volatile ", "").strip().rstrip("*").strip()

        node_names = []
        mod = module or (parsed.get("module") or "unknown").upper()
        sandbox.vectors.remove_by_metadata(source_file=filename)

        # ── Collections for two-pass derived-relationship creation ────────
        struct_node_ids: dict = {}    # struct_name → canonical node_id
        typedef_node_ids: dict = {}   # typedef_name → canonical node_id
        enum_node_ids: dict = {}      # enum_name → canonical node_id
        param_info: list = []         # (param_node_id, raw_type, func_node_id)
        member_info: list = []        # (member_node_id, raw_type, struct_node_id)
        ret_type_seen: set = set()    # dedup ReturnType nodes
        func_node_ids_by_name: dict = {}  # func_name → canonical node_id

        # ── Functions ─────────────────────────────────────────────────────
        for fn in parsed.get("functions", []):
            if not isinstance(fn, dict):
                continue
            name = fn.get("name", "")
            if not name:
                continue

            return_type = fn.get("return_type", "")
            params_list = fn.get("parameters", [])
            req_ids = fn.get("trace_info", {}).get("requirements", [])

            # signature_hash — matches production _signature_hash()
            norm_params = [
                p.get("type", "").strip()
                for p in params_list if isinstance(p, dict)
            ]
            sig = f"{name}|{return_type.strip()}|{','.join(norm_params)}"
            sig_hash = hashlib.sha1(sig.encode("utf-8")).hexdigest()[:12]

            node_type = "Function"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": name, "module": mod})
            func_node_ids_by_name[name] = node_id

            props = {
                "name": name,
                "id": f"FUNC_{name}",
                "label": name,
                "brief": fn.get("brief", ""),
                "purpose": fn.get("purpose") or fn.get("detailed_description", ""),
                "return_type": return_type,
                "return_brief": fn.get("return_details", "") or fn.get("return_doc", ""),
                "signature_hash": sig_hash,
                "traces": json.dumps(req_ids) if req_ids else None,
                "origin": "arch_only",
                "source_files": [filename],
                "source": "SWA_Functions",
                "module": mod,
                "source_file": filename,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props,
                                       clear_rel_types=self.ILLD_DETECTABLE_REL_TYPES)
            node_names.append(name)

            # RAG chunk — enriched with parameter info (H2)
            brief = fn.get("brief", "")
            params_text = ", ".join(
                f"{p.get('type', '')} {p.get('name', '')}".strip()
                for p in params_list if isinstance(p, dict)
            )
            text = (
                f"Function: {name}\nReturn: {return_type or 'void'}\nBrief: {brief}"
                + (f"\nParameters: {params_text}" if params_text else "")
            )
            sandbox.vectors.add_chunks([Chunk(
                text=text,
                source_file=filename,
                metadata={"source_file": filename, "node_type": "Function",
                          "node_id": node_id, "module": mod},
            )])

            # ReturnType node + RETURN_TYPE edge (dedup across functions)
            if return_type and return_type != "void":
                rt_id = f"RETTYPE_{return_type}"
                rt_node_id = EphemeralGraph._canonical_id(
                    "ReturnType", {"name": rt_id, "module": mod})
                if rt_id not in ret_type_seen:
                    ret_type_seen.add(rt_id)
                    self._add_node_with_shadow(sandbox, "ReturnType", rt_node_id, {
                        "id": rt_id, "name": return_type, "label": return_type, "module": mod,
                    })
                sandbox.graph.add_relationship(
                    node_id, rt_node_id, "RETURN_TYPE", {"_origin": "sandbox"})

            # DEPENDS_ON edges (with dependency_type attr matching prod)
            for dep in fn.get("dependencies", []):
                if dep and isinstance(dep, str):
                    dep_node_id = EphemeralGraph._canonical_id(
                        "Function", {"name": dep, "module": mod})
                    sandbox.graph.add_relationship(
                        node_id, dep_node_id, "DEPENDS_ON",
                        {"_origin": "sandbox", "dependency_type": "Calls"})

            # Parameter nodes + HAS_PARAMETER edges (with position attr)
            for idx, param in enumerate(params_list):
                if not isinstance(param, dict):
                    continue
                pname = param.get("name", "")
                if not pname:
                    continue
                ptype = param.get("type", "")
                p_node_type = "Parameter"
                pid_key = f"{name}__{pname}"               # canonical key (unique)
                pid_prod = f"PARAM_{name}_{pname}"         # prod-matching id property
                p_node_id = EphemeralGraph._canonical_id(
                    p_node_type, {"name": pid_key, "module": mod})
                p_props = {
                    "name": pname,
                    "id": pid_prod,
                    "label": f"{pname}: {ptype}",
                    "type": ptype,
                    "module": mod,
                    "source_file": filename,
                }
                self._add_node_with_shadow(sandbox, p_node_type, p_node_id, p_props)
                sandbox.graph.add_relationship(
                    node_id, p_node_id, "HAS_PARAMETER",
                    {"_origin": "sandbox", "position": idx})
                param_info.append((p_node_id, ptype, node_id))

            # IMPLEMENTED_BY + IMPLEMENTS edges from \trace doxygen tags (C1)
            for req_id in req_ids:
                if not req_id:
                    continue
                req_node_type = "Requirement"
                req_node_id = EphemeralGraph._canonical_id(
                    req_node_type, {"name": req_id, "module": mod})
                if sandbox.graph.get_node(req_node_id) is None:
                    sandbox.graph.add_node(req_node_type, req_node_id, {
                        "requirement_id": req_id, "name": req_id, "module": mod,
                        "source": "SWA_trace_tag", "source_file": filename,
                    })
                sandbox.graph.add_relationship(
                    req_node_id, node_id, "IMPLEMENTED_BY", {"_origin": "sandbox"})
                sandbox.graph.add_relationship(
                    node_id, req_node_id, "IMPLEMENTS", {"_origin": "sandbox"})

        # ── Structs + StructMembers ───────────────────────────────────────
        for s in parsed.get("structs", []):
            if not isinstance(s, dict):
                continue
            sname = s.get("name", "")
            if not sname:
                continue
            node_type = "Struct"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": sname, "module": mod})
            struct_node_ids[sname] = node_id
            props = {
                "name": sname,
                "id": f"STRUCT_{sname}",
                "label": sname,
                "brief": s.get("brief", ""),
                "purpose": "",   # LLM-only field — empty without enrichment pipeline
                "source": "SWA_Structs",
                "module": mod,
                "source_file": filename,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props)
            node_names.append(sname)
            members_text = ", ".join(
                m.get("name", "") for m in s.get("members", []) if isinstance(m, dict))
            text = f"Struct: {sname}\nBrief: {s.get('brief', '')}\nMembers: {members_text}"
            sandbox.vectors.add_chunks([Chunk(
                text=text, source_file=filename,
                metadata={"source_file": filename, "node_type": "Struct",
                          "node_id": node_id, "module": mod},
            )])

            for m in s.get("members", []):
                if not isinstance(m, dict):
                    continue
                mname = m.get("name", "")
                if not mname:
                    continue
                mtype = m.get("type", "unknown")
                m_node_type = "StructMember"
                mid = f"MEMBER_{sname}_{mname}"
                m_node_id = EphemeralGraph._canonical_id(m_node_type, {"name": mid, "module": mod})
                m_props = {
                    "name": mname,
                    "id": mid,
                    "label": f"{mname}: {mtype}",
                    "type": mtype,
                    "description": m.get("description", ""),
                    "module": mod,
                    "source_file": filename,
                }
                self._add_node_with_shadow(sandbox, m_node_type, m_node_id, m_props)
                sandbox.graph.add_relationship(
                    node_id, m_node_id, "HAS_MEMBER", {"_origin": "sandbox"})
                member_info.append((m_node_id, mtype, node_id))

        # ── Enums + EnumValues ────────────────────────────────────────────
        for e in parsed.get("enums", []):
            if not isinstance(e, dict):
                continue
            ename = e.get("name", "")
            if not ename:
                continue
            node_type = "Enum"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": ename, "module": mod})
            enum_node_ids[ename] = node_id
            props = {
                "name": ename,
                "id": f"ENUM_{ename}",
                "label": ename,
                "brief": e.get("brief", ""),
                "purpose": "",   # LLM-only
                "source": "SWA_Enums",
                "module": mod,
                "source_file": filename,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props)
            node_names.append(ename)
            vals_text = ", ".join(
                v.get("name", "") for v in e.get("values", []) if isinstance(v, dict))
            text = f"Enum: {ename}\nBrief: {e.get('brief', '')}\nValues: {vals_text}"
            sandbox.vectors.add_chunks([Chunk(
                text=text, source_file=filename,
                metadata={"source_file": filename, "node_type": "Enum",
                          "node_id": node_id, "module": mod},
            )])

            for v in e.get("values", []):
                if not isinstance(v, dict):
                    continue
                vname = v.get("name", "")
                if not vname:
                    continue
                v_node_type = "EnumValue"
                vid = f"ENUMVAL_{vname}"
                v_node_id = EphemeralGraph._canonical_id(v_node_type, {"name": vid, "module": mod})
                v_props = {
                    "name": vname,
                    "id": vid,
                    "label": vname,
                    "value": v.get("value", ""),
                    "description": v.get("description", ""),
                    "module": mod,
                    "source_file": filename,
                }
                self._add_node_with_shadow(sandbox, v_node_type, v_node_id, v_props)
                sandbox.graph.add_relationship(
                    node_id, v_node_id, "HAS_VALUE", {"_origin": "sandbox"})

        # ── Typedefs ──────────────────────────────────────────────────────
        for td in parsed.get("typedefs", []):
            if not isinstance(td, dict):
                continue
            tname = td.get("name", "")
            if not tname:
                continue
            node_type = "Typedef"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": tname, "module": mod})
            typedef_node_ids[tname] = node_id
            props = {
                "name": tname,
                "id": f"TYPEDEF_{tname}",
                "label": tname,
                "brief": td.get("brief", ""),
                "purpose": "",   # LLM-only
                "underlying_type": td.get("type", ""),
                "source": "SWA_Typedefs",
                "module": mod,
                "source_file": filename,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props)
            node_names.append(tname)
            text = (f"Typedef: {tname}\nUnderlying type: {td.get('type', '')}"
                    f"\nBrief: {td.get('brief', '')}")
            sandbox.vectors.add_chunks([Chunk(
                text=text, source_file=filename,
                metadata={"source_file": filename, "node_type": "Typedef",
                          "node_id": node_id, "module": mod},
            )])

            # PrimitiveType node + ALIASES edge
            utype_clean = _clean_type(td.get("type", ""))
            if utype_clean.lower() in {p.lower() for p in _PRIMITIVE_TYPES}:
                pt_id = f"PRIMITIVE_{utype_clean}"
                pt_node_id = EphemeralGraph._canonical_id(
                    "PrimitiveType", {"name": pt_id, "module": mod})
                if sandbox.graph.get_node(pt_node_id) is None:
                    sandbox.graph.add_node("PrimitiveType", pt_node_id, {
                        "id": pt_id, "name": utype_clean,
                        "label": utype_clean, "module": mod,
                    })
                sandbox.graph.add_relationship(
                    node_id, pt_node_id, "ALIASES", {"_origin": "sandbox"})

        # ── Macros ────────────────────────────────────────────────────────
        for macro in parsed.get("macros", []):
            if not isinstance(macro, dict):
                continue
            mname = macro.get("name", "")
            if not mname:
                continue
            node_type = "Macro"
            node_id = EphemeralGraph._canonical_id(node_type, {"name": mname, "module": mod})
            props = {
                "name": mname,
                "id": f"MACRO_{mname}",
                "label": mname,
                "value": macro.get("value", ""),
                "description": macro.get("description", ""),
                "purpose": "",   # LLM-only
                "source": "SWA_Macros",
                "module": mod,
                "source_file": filename,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props)
            node_names.append(mname)
            text = (f"Macro: {mname}\nValue: {macro.get('value', '')}"
                    f"\nDescription: {macro.get('description', '')}")
            sandbox.vectors.add_chunks([Chunk(
                text=text, source_file=filename,
                metadata={"source_file": filename, "node_type": "Macro",
                          "node_id": node_id, "module": mod},
            )])

        # ── Pass 2: Derived relationships (type resolution) ───────────────
        struct_names  = set(struct_node_ids)
        typedef_names = set(typedef_node_ids)
        enum_names    = set(enum_node_ids)

        # OF_TYPE: Parameter → Struct / Typedef / Enum
        # USED_BY: Struct → Function (derived from param→struct OF_TYPE)
        for p_node_id, ptype_raw, func_node_id in param_info:
            clean = _clean_type(ptype_raw)
            if clean in struct_names:
                tgt = struct_node_ids[clean]
                sandbox.graph.add_relationship(
                    p_node_id, tgt, "OF_TYPE", {"_origin": "sandbox"})
                sandbox.graph.add_relationship(
                    tgt, func_node_id, "USED_BY",
                    {"_origin": "sandbox", "usage_context": "Parameter"})
            elif clean in typedef_names:
                sandbox.graph.add_relationship(
                    p_node_id, typedef_node_ids[clean], "OF_TYPE", {"_origin": "sandbox"})
            elif clean in enum_names:
                sandbox.graph.add_relationship(
                    p_node_id, enum_node_ids[clean], "OF_TYPE", {"_origin": "sandbox"})

        # OF_TYPE: StructMember → Typedef
        # USED_IN: Typedef → Struct (derived from member→typedef OF_TYPE)
        for m_node_id, mtype_raw, struct_node_id in member_info:
            clean = _clean_type(mtype_raw)
            if clean in typedef_names:
                tgt = typedef_node_ids[clean]
                sandbox.graph.add_relationship(
                    m_node_id, tgt, "OF_TYPE", {"_origin": "sandbox"})
                sandbox.graph.add_relationship(
                    tgt, struct_node_id, "USED_IN",
                    {"_origin": "sandbox", "member_name": ""})

        return node_names

    def _ingest_sfr(self, sandbox, parsed, filename, module) -> List[str]:
        """Ingest SFR parser output → HardwareRegister + RegisterField nodes + HAS_FIELD edges.

        Matches production node types from illd_kg_builder.py (v3.0 collapsed labels).
        Production IDs: HWREG_{MODULE}_{reg_name}, REGFIELD_{MODULE}_{reg_name}_{field_name}
        """
        node_names = []
        # Canonicalize submodule regdef tokens (e.g. PMSCORE/PMSMON → PMS).
        mod = self._canonicalize_illd_module(module or parsed.get("module"), filename)
        registers_dict = parsed.get("registers", {})
        sandbox.vectors.remove_by_metadata(source_file=filename)

        hw_reg_count = 0
        reg_field_count = 0

        for reg_name, bitfields in registers_dict.items():
            if not isinstance(bitfields, list):
                continue

            # HardwareRegister node — matches prod ID convention
            node_type = "HardwareRegister"
            reg_id = f"HWREG_{mod}_{reg_name}"
            node_id = EphemeralGraph._canonical_id(
                node_type, {"name": reg_name, "module": mod})
            props = {
                "id": reg_id,
                "name": reg_name,
                "module": mod,
                "label": reg_name,
                "sfr_source_file": filename,
                "source_file": filename,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props)
            node_names.append(reg_name)
            hw_reg_count += 1

            # Semantic chunk for the register
            bf_text = ", ".join(bf.get("name", "") for bf in bitfields if isinstance(bf, dict))
            text = f"HardwareRegister: {reg_name}\nModule: {mod}\nFields: {bf_text}"
            sandbox.vectors.add_chunks([Chunk(
                text=text,
                source_file=filename,
                metadata={"source_file": filename, "node_type": "HardwareRegister",
                          "node_id": node_id, "module": mod},
            )])

            # RegisterField nodes + HAS_FIELD edges
            for bf in bitfields:
                if not isinstance(bf, dict):
                    continue
                bfname = bf.get("name", "")
                if not bfname:
                    continue
                field_node_type = "RegisterField"
                field_id = f"REGFIELD_{mod}_{reg_name}_{bfname}"
                field_node_id = EphemeralGraph._canonical_id(
                    field_node_type, {"name": field_id, "module": mod})
                field_props = {
                    "name": bfname,
                    "id": field_id,
                    "register": reg_name,
                    "module": mod,
                    "bit_range": bf.get("bit_range", ""),
                    "bits": bf.get("bit_range", "") or bf.get("bits", ""),
                    "width": bf.get("width", ""),
                    "reset_value": bf.get("reset_value", ""),
                    "description": bf.get("description", ""),
                    "sfr_source_file": filename,
                    "source_file": filename,
                    "label": bf.get("label", "") or f"{bfname} {bf.get('bit_range', '')}",
                }
                if bf.get("access_type"):
                    field_props["access_type"] = bf["access_type"]
                self._add_node_with_shadow(sandbox, field_node_type, field_node_id, field_props)
                sandbox.graph.add_relationship(
                    node_id, field_node_id, "HAS_FIELD", {"_origin": "sandbox"})
                reg_field_count += 1

        logger.info(
            "[SandboxAdapter] SFR ingestion: %d HardwareRegisters, %d RegisterFields, "
            "file=%s, module=%s",
            hw_reg_count, reg_field_count, filename, mod,
        )
        return node_names

    @staticmethod
    def _canonicalize_illd_module(module: Optional[str], filename: Optional[str] = None) -> str:
        """Resolve regdef submodule tokens to canonical iLLD module names."""
        global _ILLD_MODULE_ALIAS_CACHE
        if _ILLD_MODULE_ALIAS_CACHE is None:
            aliases: Dict[str, str] = {
                "LPBTM": "BTM",
                "LPCAN": "CAN",
            }
            try:
                cfg_path = Path(__file__).resolve().parents[2] / "HybridRAG" / "config" / "illd_module_map.yaml"
                raw = _yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
                modules = raw.get("modules", raw) if isinstance(raw, dict) else {}
                for module_name, info in modules.items():
                    if not isinstance(info, dict):
                        continue
                    canonical = str(module_name).upper()
                    aliases[canonical] = canonical
                    for sfr_name in info.get("sfr") or []:
                        if not isinstance(sfr_name, str):
                            continue
                        stem = Path(sfr_name).stem
                        aliases[stem.lower()] = canonical
                        match = re.search(r"Ifx(\w+)_regdef", stem, re.IGNORECASE)
                        if match:
                            aliases[match.group(1).upper()] = canonical
            except Exception:
                pass
            _ILLD_MODULE_ALIAS_CACHE = aliases

        aliases = _ILLD_MODULE_ALIAS_CACHE or {}
        if module:
            token = str(module).strip()
            if token:
                return aliases.get(token.upper(), token.upper())
        if filename:
            stem = Path(filename).stem
            if stem.lower() in aliases:
                return aliases[stem.lower()]
            match = re.search(r"Ifx(\w+)_regdef", stem, re.IGNORECASE)
            if match:
                token = match.group(1).upper()
                return aliases.get(token, token)
        return "unknown"

    def _ingest_xlsx_requirements_illd(self, sandbox, parsed, filename, module) -> List[str]:
        """Ingest iLLD requirements from xlsx Jama export.

        Expected parsed format (from _parse_xlsx detecting iLLD Jama pattern):
          {"type": "illd_xlsx_req", "requirements": [...], "module": str}
        """
        node_names = []
        mod = module or parsed.get("module", "unknown")
        sandbox.vectors.remove_by_metadata(source_file=filename)

        for req in parsed.get("requirements", []):
            if not isinstance(req, dict):
                continue
            req_id = req.get("requirement_id", "")
            if not req_id:
                continue
            node_type = "Requirement"
            node_id = EphemeralGraph._canonical_id(
                node_type, {"name": req_id, "module": mod})
            req_desc = req.get("description", "")
            props = {
                "requirement_id": req_id,
                "name": req.get("name", req_id),
                "description": req_desc,
                "module": mod,
                "status": req.get("status", ""),
                "source_file": filename,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props)
            node_names.append(req_id)
            # Semantic chunk — include description when available (M1)
            text = (
                f"Requirement: {req_id}\nName: {req.get('name', '')}"
                + (f"\nDescription: {req_desc[:300]}" if req_desc else "")
            )
            sandbox.vectors.add_chunks([Chunk(
                text=text,
                source_file=filename,
                metadata={"source_file": filename, "node_type": "Requirement",
                          "node_id": node_id, "module": mod},
            )])

        return node_names

    def _ingest_requirements_illd(self, sandbox, parsed, filename, module) -> List[str]:
        """Ingest iLLD requirements from JSON (Jama API export format).

        Expected parsed format:
          {"type": "json", "data": {"requirements": [...]}}
        """
        node_names = []
        mod = module or "unknown"
        data = parsed.get("data", {})
        reqs = data.get("requirements") if isinstance(data, dict) else None
        if not isinstance(reqs, list):
            return node_names

        sandbox.vectors.remove_by_metadata(source_file=filename)
        for req in reqs:
            if not isinstance(req, dict):
                continue
            req_id = (req.get("document_key") or req.get("requirement_id")
                      or req.get("id") or "")
            if not req_id:
                continue
            node_type = "Requirement"
            node_id = EphemeralGraph._canonical_id(
                node_type, {"name": req_id, "module": mod})
            props = {
                "requirement_id": req_id,
                "name": req.get("name", req_id),
                "description": req.get("description", ""),
                "module": mod,
                "status": str(req.get("status", "")),
                "source_file": filename,
            }
            self._add_node_with_shadow(sandbox, node_type, node_id, props)
            node_names.append(req_id)
            # Semantic chunk
            desc = req.get("description", "")
            text = f"Requirement: {req_id}\nName: {req.get('name', '')}\nDescription: {desc[:300]}"
            sandbox.vectors.add_chunks([Chunk(
                text=text,
                source_file=filename,
                metadata={"source_file": filename, "node_type": "Requirement",
                          "node_id": node_id, "module": mod},
            )])

        return node_names

    # ── Shadow/Override Logic (Phase 5) ────────────────────────────────

    # Relationship types the sandbox parser can reliably detect even
    # without full include paths.  When a node is shadowed, prod edges
    # of these types are cleared so the sandbox parser re-creates only
    # what actually exists in the developer's modified code.  Prod edges
    # for types NOT in this set (SRC_IMPLEMENTS_EA, etc.) are inherited.
    def _remove_deleted_prod_nodes(self, sandbox: 'EphemeralSandbox',
                                    filename: str) -> int:
        """Remove prod nodes from *filename* that were not re-created by sandbox.

        After ingesting a file, any node whose ``source_file`` matches
        *filename* and whose ``_origin`` is still ``'production'`` was never
        shadowed — meaning the developer removed it.  We drop it from the
        graph so the sandbox reflects reality.

        Vectors are already clean: every ``_ingest_*_illd`` method starts with
        ``sandbox.vectors.remove_by_metadata(source_file=filename)`` and only
        re-adds chunks for entities present in the new file.

        Returns the number of nodes removed.
        """
        fname_base = Path(filename).name
        to_delete = [
            nid
            for nid, data in list(sandbox.graph.get_all_nodes())
            if data.get("_origin") == "production"
            and Path(data.get("source_file", "")).name == fname_base
        ]
        for nid in to_delete:
            sandbox.graph.remove_node(nid)
        return len(to_delete)

    SANDBOX_DETECTABLE_REL_TYPES = {"SRC_CALLS"}

    # Extended set: when include_paths are provided, clang can also detect
    # SFR accesses and global variable references reliably.
    SANDBOX_DETECTABLE_REL_TYPES_WITH_INCLUDES = {
        "SRC_CALLS", "SRC_ACCESSES_SFR", "SRC_USES_GLOBAL",
    }

    # iLLD-specific: call-graph edges that the sandbox can reliably detect
    ILLD_DETECTABLE_REL_TYPES = {"CALLS_INTERNALLY"}

    def _add_node_with_shadow(self, sandbox: 'EphemeralSandbox', node_type: str,
                              node_id: str, properties: Dict,
                              clear_rel_types: Optional[set] = None):
        """Add node; if it shadows a prod node, mark the override.

        Parameters
        ----------
        clear_rel_types : set | None
            Outgoing prod edge types to remove from the shadowed node.
            Defaults to SANDBOX_DETECTABLE_REL_TYPES.  These are edge
            types the sandbox parser will re-create from the uploaded
            source, so the stale prod versions must be cleared first.
        """
        existing = sandbox.graph.get_node(node_id)
        if existing and existing.get("_origin") == "production":
            # Shadow: replace prod node with sandbox version
            properties["_origin"] = "sandbox"
            properties["_shadows"] = node_id
            properties["_original_prod_properties"] = {
                k: v for k, v in existing.items() if not k.startswith("_")
            }
            # Clear stale prod edges that sandbox will re-detect
            types_to_clear = clear_rel_types if clear_rel_types is not None else self.SANDBOX_DETECTABLE_REL_TYPES
            if types_to_clear:
                sandbox.graph.remove_outgoing_edges_by_type(node_id, types_to_clear)
            sandbox.graph.update_node(node_id, properties)
        else:
            properties["_origin"] = "sandbox"
            sandbox.graph.add_node(node_type, node_id, properties)

    # ── Helpers ────────────────────────────────────────────────────────

    @staticmethod
    def _iter_c_functions(parsed: Dict):
        """Yield function dicts from C parser output.

        Handles both formats:
        - dict:  {"func_name": {"parameters": ..., ...}, ...}  (regex/clang parser)
        - list:  [{"name": "func_name", ...}, ...]              (fallback parser)
        """
        functions = parsed.get("functions", [])
        if isinstance(functions, dict):
            for name, data in functions.items():
                fn = {"name": name}
                if isinstance(data, dict):
                    fn.update(data)
                yield fn
        elif isinstance(functions, list):
            for fn in functions:
                if isinstance(fn, dict) and fn.get("name"):
                    yield fn

    @staticmethod
    def _extract_text_content(parsed: Dict) -> str:
        """Extract text content from any parsed type for vector chunking."""
        ptype = parsed.get("type", "")
        if ptype in ("c_source", "c_header"):
            return parsed.get("content", "")
        elif ptype == "pdf_md":
            return parsed.get("raw_md", "")
        elif ptype == "pdf":
            return "\n".join(parsed.get("pages", []))
        elif ptype == "rst":
            sections = parsed.get("sections", [])
            if isinstance(sections, list):
                return "\n".join(str(s) for s in sections)
            return str(sections)
        elif ptype == "json":
            import json as _json
            return _json.dumps(parsed.get("data", {}), indent=2, default=str)[:50000]
        elif ptype in ("text", "generic"):
            return parsed.get("content", "")
        return ""

    def _chunk_text(self, content: str, filename: str) -> List[Chunk]:
        """Split content into overlapping chunks for vector embedding."""
        chunks = []
        lines = content.split("\n")
        buf, section = [], ""
        for line in lines:
            if re.match(r'^(#{1,3}|/\*\*|\*\s+@|//\s*={3,})', line):
                if buf:
                    chunks.append(Chunk(
                        text="\n".join(buf), source_file=filename, section=section,
                    ))
                    buf = []
                section = line.strip()[:80]
            buf.append(line)
            if len("\n".join(buf)) > self._chunk_size:
                chunks.append(Chunk(text="\n".join(buf), source_file=filename, section=section))
                buf = []
        if buf:
            chunks.append(Chunk(text="\n".join(buf), source_file=filename, section=section))
        return chunks


# ═════════════════════════════════════════════════════════════════════════
#  TraceabilityPuller — Plan 2 Phase 4
#  Pull ±N traceability neighbors from prod Neo4j
# ═════════════════════════════════════════════════════════════════════════

class TraceabilityPuller:
    """Pull ±N traceability neighbors from prod Neo4j for given node names.

    Read-only: no writes, no locks. Returns (nodes, relationships).
    """

    MAX_PULL_NODES = 500  # safety cap per upload call

    def __init__(self, neo4j_driver):
        self._driver = neo4j_driver

    @staticmethod
    def _resolve_database(workspace_id: str) -> str:
        """Resolve the actual Neo4j database name from storage_config.yaml.

        The workspace_id (e.g. 'mcal', 'illd') is NOT the database name —
        both profiles use database='neo4j' in the config.
        """
        try:
            from src.HybridRAG.code.neo4j_manager import get_instance_config
            cfg = get_instance_config(workspace_id)
            return cfg.database or "neo4j"
        except Exception:
            return "neo4j"

    def pull_neighbors(
        self, node_names: List[str], module: str,
        workspace_id: str = "mcal", depth: int = 1,
    ) -> tuple:
        """Pull nodes ±depth hops from the named nodes in prod KG.

        Returns (nodes: List[dict], relationships: List[dict]).
        Each node dict: {"node_id": str, "node_type": str, "properties": dict}
        Each rel dict: {"source": str, "target": str, "rel_type": str, "properties": dict}
        """
        if not node_names or depth <= 0:
            return [], []

        db = self._resolve_database(workspace_id)
        module_upper = module.upper() if module else ""
        # depth and max_nodes are injected as literals because Neo4j 4.x
        # does not allow parameters inside [*1..N] or list slices [0..N].
        # elementId() replaced with toString(id()) for Neo4j 4.x compat.
        # F-CB-01: bound depth to 1..5 as defense-in-depth before literal
        # interpolation. MCP layer already validates trace_depth in 0..2,
        # but callers from other paths must not be able to inject Cypher
        # via crafted depth values or run unbounded variable-length paths.
        safe_depth = int(depth)
        if not 1 <= safe_depth <= 5:
            raise ValueError(
                f"TraceabilityPuller.pull_neighbors: depth must be 1..5, got {depth!r}"
            )
        safe_max = int(self.MAX_PULL_NODES)
        cypher = f"""
        MATCH (n)
        WHERE toUpper(n.module) = $module_upper
          AND (n.name IN $names OR n.function_name IN $names
               OR n.param_name IN $names OR n.requirement_id IN $names)
        WITH n LIMIT 200
        CALL {{
            WITH n
            MATCH path = (n)-[*1..{safe_depth}]-(neighbor)
            WHERE toUpper(neighbor.module) = $module_upper
            RETURN neighbor, relationships(path) AS rels
        }}
        WITH collect(DISTINCT n) + collect(DISTINCT neighbor) AS all_nodes,
             collect(rels) AS all_rel_lists
        UNWIND all_nodes AS node
        WITH collect(DISTINCT {{
                 node_id: toString(id(node)),
                 node_type: labels(node)[0],
                 properties: properties(node)
             }})[0..{safe_max}] AS nodes,
             all_rel_lists
        UNWIND all_rel_lists AS rel_list
        UNWIND rel_list AS rel
        WITH nodes, collect(DISTINCT {{
                 source: toString(id(startNode(rel))),
                 target: toString(id(endNode(rel))),
                 rel_type: type(rel),
                 properties: properties(rel)
             }}) AS relationships
        RETURN nodes, relationships
        """
        try:
            with self._driver.session(database=db) as session:
                result = session.run(
                    cypher,
                    names=node_names,
                    module_upper=module_upper,
                )
                record = result.single()
                if record:
                    nodes = record["nodes"] or []
                    rels = record["relationships"] or []
                    if len(nodes) >= self.MAX_PULL_NODES:
                        logger.warning(
                            "[TraceabilityPuller] Hit %d-node cap for module=%s, depth=%d",
                            self.MAX_PULL_NODES, module, depth,
                        )
                    return nodes, rels
                return [], []
        except Exception as e:
            logger.error("[TraceabilityPuller] Failed to pull neighbors: %s", e)
            return [], []

    def pull_boundary_nodes(
        self, names: List[str], workspace_id: str = "mcal",
    ) -> List[Dict]:
        """Resolve boundary node names against prod KG (cross-module).

        Unlike pull_neighbors(), this does NOT filter by module — it finds
        nodes by name across all modules. Returns only the anchor nodes
        (no neighbor expansion) for lightweight boundary resolution.

        Returns list of node dicts: {"node_id": str, "node_type": str, "properties": dict}
        """
        if not names:
            return []

        db = self._resolve_database(workspace_id)
        # Neo4j 4.x: elementId() → toString(id()), literal max_nodes
        safe_max = int(self.MAX_PULL_NODES)
        cypher = f"""
        MATCH (n)
        WHERE n.name IN $names OR n.function_name IN $names
        RETURN collect(DISTINCT {{
                   node_id: toString(id(n)),
                   node_type: labels(n)[0],
                   properties: properties(n)
               }})[0..{safe_max}] AS nodes
        """
        try:
            with self._driver.session(database=db) as session:
                result = session.run(
                    cypher, names=names,
                )
                record = result.single()
                if record:
                    nodes = record["nodes"] or []
                    logger.info(
                        "[TraceabilityPuller] Boundary resolution: %d/%d names resolved",
                        len(nodes), len(names),
                    )
                    return nodes
                return []
        except Exception as e:
            logger.error("[TraceabilityPuller] Failed boundary pull: %s", e)
            return []


# ═════════════════════════════════════════════════════════════════════════
#  HybridGraphService — Plan 2 Phase 6
#  Route queries to sandbox or Neo4j based on traversal depth
# ═════════════════════════════════════════════════════════════════════════

class HybridGraphService:
    """Route queries to sandbox or Neo4j based on traversal depth needed."""

    # Shallow tools → sandbox NetworkX (has user's changes + ±1 context)
    SHALLOW_TOOLS = {
        "search_database", "search_nodes", "get_node_by_id",
        "get_neighbors", "sandbox_query", "sandbox_status",
        "query_api_function", "get_type_definition",
        "query_dependencies", "get_distribution",
    }

    # Deep tools → prod Neo4j + patch with sandbox overrides
    DEEP_TOOLS = {
        "find_coverage_gaps", "build_traceability_matrix",
        "find_requirement_traces", "shortest_path",
        "analyze_hw_sw_links", "detect_communities",
        "get_ontology_compliance", "get_coverage_report",
        "get_graph_statistics", "get_failure_patterns",
        "get_review_analytics", "get_learning_metrics",
        "execute_cypher",
    }

    def __init__(self, sandbox: EphemeralSandbox, neo4j_driver=None,
                 qdrant_client=None, workspace_id: str = "mcal"):
        self._sandbox = sandbox
        self._driver = neo4j_driver
        self._qdrant = qdrant_client
        self._workspace_id = workspace_id

    def search(self, query: str, top_k: int = 15, alpha: float = _DEFAULT_SEARCH_ALPHA,
               filter_by_module: Optional[str] = None) -> List[SearchResult]:
        """Hybrid query — sandbox graph + vectors + production Qdrant.

        Merges sandbox results with production Qdrant results.
        Sandbox results always take priority. Prod Qdrant results from files
        that were uploaded to the sandbox are EXCLUDED (shadowed).
        """
        # Sandbox results (graph keyword + ephemeral vectors)
        querier = SandboxQuerier(self._sandbox)
        sandbox_results = querier.search(query, top_k=top_k, alpha=alpha)

        # Collect filenames uploaded to sandbox — these shadow prod versions
        sandbox_files = set()
        for f in self._sandbox.files_ingested:
            if isinstance(f, dict) and f.get("filename"):
                sandbox_files.add(f["filename"].lower())

        # Production Qdrant results (G1 fix) — excluding shadowed files
        prod_results = self._query_prod_qdrant(query, top_k=top_k,
                                                filter_by_module=filter_by_module,
                                                exclude_files=sandbox_files)

        # Merge: sandbox results take priority (higher boost for user's own data)
        seen = {}
        for r in sandbox_results:
            seen[r.node_id] = r
        for r in prod_results:
            if r.node_id not in seen:
                seen[r.node_id] = r

        merged = sorted(seen.values(), key=lambda x: -x.score)
        return merged[:top_k]

    def _query_prod_qdrant(self, query: str, top_k: int = 10,
                           filter_by_module: Optional[str] = None,
                           exclude_files: Optional[set] = None) -> List[SearchResult]:
        """Query production Qdrant for additional context not in the sandbox.

        Results from files in `exclude_files` are filtered out (those files
        have been re-uploaded to the sandbox and their prod versions are stale).
        """
        if not self._qdrant:
            return []
        exclude_files = exclude_files or set()
        try:
            # Use the same embedder as the sandbox vectors
            embedder = self._sandbox.vectors._embedder
            embedding = embedder.embed([query])
            if not embedding or not embedding[0]:
                return []
            q_vec = embedding[0]

            # Resolve collection(s) — iLLD uses module name, MCAL uses sub-collections
            module = filter_by_module
            if not module:
                # Try to infer module from sandbox context
                files = self._sandbox.files_ingested
                for f in files:
                    if isinstance(f, dict) and f.get("filename"):
                        # Module is typically uppercase in sandbox
                        pass
            collections = self._resolve_qdrant_collections(module)
            if not collections:
                return []

            results = []
            for collection in collections[:5]:  # Limit to 5 collections max
                try:
                    response = self._qdrant.query_points(
                        collection_name=collection,
                        query=q_vec,
                        limit=top_k,
                        with_payload=True,
                    )
                    hits = response.points if hasattr(response, 'points') else response
                    for hit in hits:
                        payload = hit.payload or {}

                        # Shadow filter: skip chunks from files uploaded to sandbox
                        source_file = (payload.get("source_file")
                                       or payload.get("file")
                                       or payload.get("filename") or "")
                        if source_file:
                            # Compare basename (prod may store full path)
                            basename = source_file.rsplit("/", 1)[-1].rsplit("\\", 1)[-1].lower()
                            if basename in exclude_files:
                                continue

                        node_id = (payload.get("_original_id")
                                   or payload.get("document_id")
                                   or payload.get("name")
                                   or str(hit.id))
                        content = (payload.get("document")
                                   or payload.get("text")
                                   or payload.get("content") or "")
                        node_type = (payload.get("node_type")
                                     or payload.get("type") or "Unknown")
                        results.append(SearchResult(
                            node_id=node_id,
                            content=content[:500],
                            score=float(hit.score) * 0.9,  # Slight discount vs sandbox
                            origin="prod_qdrant",
                            metadata={"collection": collection, **payload},
                        ))
                except Exception as e:
                    logger.debug("[HybridGraphService] Qdrant collection '%s' query failed: %s",
                                 collection, e)
                    continue

            results.sort(key=lambda r: -r.score)
            return results[:top_k]
        except Exception as e:
            logger.warning("[HybridGraphService] Prod Qdrant query failed: %s", e)
            return []

    def _resolve_qdrant_collections(self, module: Optional[str]) -> List[str]:
        """Resolve which Qdrant collections to query."""
        if module:
            return [module.lower()]
        # Discover all collections for the workspace
        try:
            all_cols = self._qdrant.get_collections().collections
            names = [c.name for c in all_cols]
            if self._workspace_id == "illd":
                # iLLD: bare module names or rag_* collections
                return [n for n in names if not any(p in n for p in
                        ("_swa_", "_swud_", "_testspec_", "_jama_"))][:5]
            else:
                mcal_prefixes = ("_swa_", "_swud_", "_testspec_", "_jama_")
                return [n for n in names if any(p in n for p in mcal_prefixes)][:5]
        except Exception:
            return []

    def deep_query(self, cypher: str, params: Dict, workspace_id: str = "mcal") -> List[Dict]:
        """Deep traversal — run against prod Neo4j, patch with sandbox overrides.

        Returns list of dicts, each with _origin and optional _patched/_injected flags.
        """
        if not self._driver:
            # No Neo4j driver — execute Cypher against sandbox NetworkX graph
            logger.warning("[HybridGraphService] No Neo4j driver for deep query — sandbox-only")
            return self._execute_cypher_on_graph(cypher, params)

        db = workspace_id if workspace_id in ("illd", "mcal") else "neo4j"

        try:
            with self._driver.session(database=db) as session:
                prod_results = session.run(cypher, params).data()
        except Exception as e:
            logger.warning("[HybridGraphService] Neo4j unavailable for deep query: %s", e)
            return self._execute_cypher_on_graph(cypher, params)

        # Patch: replace any node that has a sandbox override
        patched = []
        for record in prod_results:
            canonical = self._canonical_id_from_record(record)
            if canonical:
                sandbox_node = self._sandbox.graph.get_node(canonical)
                if sandbox_node and sandbox_node.get("_origin") == "sandbox":
                    record.update({k: v for k, v in sandbox_node.items() if not k.startswith("_")})
                    record["_patched"] = True
                    record["_origin"] = "sandbox"
                else:
                    record["_origin"] = "production"
            else:
                record["_origin"] = "production"
            patched.append(record)

        # Inject sandbox-only nodes (new nodes not in prod at all)
        existing_ids = set()
        for r in patched:
            cid = self._canonical_id_from_record(r)
            if cid:
                existing_ids.add(cid)

        query_labels = set(re.findall(r'\([\w]*:([\w]+)\)', cypher or "", re.IGNORECASE))
        query_module = (
            (params or {}).get("module")
            or (params or {}).get("mod")
            or (params or {}).get("module_upper")
        )

        for nid, data in self._sandbox.graph.get_nodes_by_origin("sandbox"):
            if not self._sandbox_node_matches_query(data, query_labels, query_module):
                continue
            if nid not in existing_ids:
                injected = {k: v for k, v in data.items() if not k.startswith("_")}
                injected["_injected"] = True
                injected["_origin"] = "sandbox"
                injected["_node_id"] = nid
                patched.append(injected)

        return patched

    @staticmethod
    def _sandbox_node_matches_query(data: Dict[str, Any],
                                    query_labels: set,
                                    query_module: Optional[str]) -> bool:
        """Best-effort guard so sandbox injection respects basic Cypher shape.

        Arbitrary Cypher cannot be perfectly re-evaluated in Python, but we can
        safely require at least a label or module match before injecting
        sandbox-only nodes. This prevents unrelated nodes from appearing in
        execute_cypher results.
        """
        node_type = data.get("_node_type", "")
        node_module = str(data.get("module", ""))
        if query_labels and node_type not in query_labels:
            return False
        if query_module and node_module.upper() != str(query_module).upper():
            return False
        return bool(query_labels or query_module)

    def _sandbox_only_results(self, warning: str = None) -> List[Dict]:
        """Return sandbox nodes as fallback when Neo4j is unavailable."""
        results = []
        for nid, data in self._sandbox.graph.get_all_nodes():
            entry = {k: v for k, v in data.items() if not k.startswith("_")}
            entry["_origin"] = data.get("_origin", "sandbox")
            entry["_node_id"] = nid
            results.append(entry)
        if warning:
            results.insert(0, {"_warning": warning})
        return results

    # ── Sandbox-local Cypher executor ─────────────────────────────────────

    def _execute_cypher_on_graph(self, cypher: str, params: Dict) -> List[Dict]:
        """Execute a simplified Cypher MATCH against the in-memory NetworkX graph.

        Supported patterns (DA fetchScopedContext / TestgenPipeline):
          • MATCH (n:Label) [WHERE …] RETURN n.prop
          • MATCH (a:L1)-[:REL]->(b:L2) [WHERE …] RETURN a.prop, b.prop
          • MATCH (a:L1)-[:REL]->(b:L2) RETURN a.prop, collect(b.prop) AS alias
          • MATCH (a:L1)-[:REL]->(b:L2) RETURN a.prop, collect({k: b.p}) AS alias
          • MATCH (a:L1)-[:REL]->(b) RETURN …, collect(CASE WHEN b.p IS NOT NULL
                                                        THEN {k: b.p} ELSE null END)
          • MATCH (a:L) OPTIONAL MATCH (a)-[:REL]->(b:L2) RETURN a.prop,
                                                            collect(b.prop) AS alias
          • MATCH (a:L1)-[:REL*min..max]->(b:L2) RETURN a.prop, b.prop
          • WHERE var.prop IN $list / [val1, val2]
        Unsupported syntax is silently skipped (returns []).
        """
        import re as _re
        if not cypher:
            return []

        # 1. Param substitution — list-aware
        for k, v in (params or {}).items():
            if isinstance(v, list):
                replacement = '[' + ', '.join(
                    f'"{x}"' if isinstance(x, str) else str(x) for x in v
                ) + ']'
            else:
                replacement = f'"{v}"' if isinstance(v, str) else str(v)
            cypher = _re.sub(rf'\${k}\b', replacement, cypher)

        # 2. LIMIT
        limit = 5000
        lm = _re.search(r'\bLIMIT\s+(\d+)', cypher, _re.IGNORECASE)
        if lm:
            limit = int(lm.group(1))

        # 3. RETURN clause (strip optional DISTINCT keyword)
        rm = _re.search(r'\bRETURN\b(.*?)(?:\bORDER\b|\bLIMIT\b|\bSKIP\b|$)',
                        cypher, _re.IGNORECASE | _re.DOTALL)
        if not rm:
            return []
        return_clause = rm.group(1).strip()
        is_distinct = False
        if _re.match(r'DISTINCT\b', return_clause, _re.IGNORECASE):
            is_distinct = True
            return_clause = _re.sub(r'^DISTINCT\s+', '', return_clause, flags=_re.IGNORECASE)
        has_collect = bool(_re.search(r'\bcollect\s*\(', return_clause, _re.IGNORECASE))

        # 4. WHERE clause (before RETURN / OPTIONAL MATCH)
        wm = _re.search(r'\bWHERE\b(.*?)(?:\bRETURN\b|\bWITH\b|\bOPTIONAL\b)',
                        cypher, _re.IGNORECASE | _re.DOTALL)
        where_clause = wm.group(1).strip() if wm else ""

        # Pre-compute return_items once (respects nested parens + AS aliases)
        return_items = self._split_return_items(return_clause)

        # 5. Variable-length path: (a:L1)-[:REL*min..max]->(b:L2)
        vl_m = _re.search(
            r'\(\s*(\w+)(?::(\w+))?\s*\)\s*-\s*\[:(\w+)\*(\d+)\.\.(\d+)\]\s*->\s*\(\s*(\w+)(?::(\w+))?\s*\)',
            cypher, _re.IGNORECASE,
        )
        if vl_m:
            return self._traverse_variable_length_rel(
                src_var=vl_m.group(1), src_label=vl_m.group(2),
                rel_type=vl_m.group(3),
                min_depth=int(vl_m.group(4)), max_depth=int(vl_m.group(5)),
                tgt_var=vl_m.group(6), tgt_label=vl_m.group(7),
                return_items=return_items, where_clause=where_clause, limit=limit,
            )

        # 6. OPTIONAL MATCH
        if _re.search(r'\bOPTIONAL\s+MATCH\b', cypher, _re.IGNORECASE):
            return self._execute_optional_match(cypher, return_clause, where_clause, limit)

        # 7. Relationship pattern: (a:L1)-[:REL]->(b:L2) or (a:L1)-[:REL]->(b)
        rel_m = _re.search(
            r'\(\s*(\w+)(?::(\w+))?\s*\)\s*-\s*\[:(\w+)\]\s*->\s*\(\s*(\w+)(?::(\w+))?\s*\)',
            cypher, _re.IGNORECASE,
        )
        if rel_m:
            src_var, src_label = rel_m.group(1), rel_m.group(2)
            rel_type = rel_m.group(3)
            tgt_var, tgt_label = rel_m.group(4), rel_m.group(5)
            if has_collect:
                rows = self._traverse_rel_grouped(
                    src_var, src_label, rel_type, tgt_var, tgt_label,
                    return_clause, where_clause, limit,
                )
            else:
                rows = self._traverse_rel(
                    src_var, src_label, rel_type, tgt_var, tgt_label,
                    return_items, where_clause, limit,
                )
            if is_distinct:
                rows = self._dedup_rows(rows)
            return rows

        # 8. Single-node pattern: (n:Label) or (n)
        node_m = _re.search(r'\(\s*(\w+)(?::(\w+))?\s*\)', cypher, _re.IGNORECASE)
        if node_m:
            rows = self._match_single_nodes(
                node_var=node_m.group(1), node_label=node_m.group(2),
                return_items=return_items, where_clause=where_clause, limit=limit,
            )
            if is_distinct:
                rows = self._dedup_rows(rows)
            return rows

        return []

    @staticmethod
    def _dedup_rows(rows: List[Dict]) -> List[Dict]:
        """Deduplicate rows for RETURN DISTINCT, preserving order."""
        seen: set = set()
        out: List[Dict] = []
        for row in rows:
            key = tuple(sorted((k, str(v)) for k, v in row.items()))
            if key not in seen:
                seen.add(key)
                out.append(row)
        return out

    def _traverse_rel(self, src_var: str, src_label: Optional[str],
                      rel_type: str,
                      tgt_var: str, tgt_label: Optional[str],
                      return_items: List[tuple], where_clause: str,
                      limit: int) -> List[Dict]:
        """Walk NetworkX edges matching (src_label)-[:rel_type]->(tgt_label)."""
        results: List[Dict] = []
        g = self._sandbox.graph._graph
        for src_id, tgt_id, edge_data in g.edges(data=True):
            if edge_data.get("_rel_type") != rel_type:
                continue
            src_data = dict(g.nodes.get(src_id, {}))
            tgt_data = dict(g.nodes.get(tgt_id, {}))
            if src_label and src_data.get("_node_type") != src_label:
                continue
            if tgt_label and tgt_data.get("_node_type") != tgt_label:
                continue
            var_map = {src_var: src_data, tgt_var: tgt_data}
            if where_clause and not self._eval_where(where_clause, var_map):
                continue
            results.append(self._build_row(return_items, var_map))
            if len(results) >= limit:
                break
        return results

    def _traverse_rel_grouped(self, src_var: str, src_label: Optional[str],
                               rel_type: str,
                               tgt_var: str, tgt_label: Optional[str],
                               return_clause: str, where_clause: str,
                               limit: int) -> List[Dict]:
        """Walk edges and apply collect() grouping from RETURN clause."""
        import re as _re
        g = self._sandbox.graph._graph
        # Build ordered map: src_id → (src_data, [tgt_data, ...])
        groups: Dict[str, tuple] = {}
        order: List[str] = []
        for src_id, tgt_id, edge_data in g.edges(data=True):
            if edge_data.get("_rel_type") != rel_type:
                continue
            src_data = dict(g.nodes.get(src_id, {}))
            tgt_data = dict(g.nodes.get(tgt_id, {}))
            if src_label and src_data.get("_node_type") != src_label:
                continue
            if tgt_label and tgt_data.get("_node_type") != tgt_label:
                continue
            if where_clause and not self._eval_where(
                    where_clause, {src_var: src_data, tgt_var: tgt_data}):
                continue
            if src_id not in groups:
                groups[src_id] = (src_data, [])
                order.append(src_id)
            groups[src_id][1].append(tgt_data)

        return_items = self._split_return_items(return_clause)
        results: List[Dict] = []
        for src_id in order:
            src_data, tgt_list = groups[src_id]
            row: Dict = {}
            for expr, alias in return_items:
                collect_m = _re.match(r'collect\s*\((.+)\)$', expr.strip(),
                                      _re.IGNORECASE | _re.DOTALL)
                if collect_m:
                    collected = []
                    for tgt_data in tgt_list:
                        val = self._eval_expr(
                            collect_m.group(1).strip(),
                            {src_var: src_data, tgt_var: tgt_data},
                        )
                        if val is not None:
                            collected.append(val)
                    row[alias] = collected
                else:
                    row[alias] = self._eval_expr(expr, {src_var: src_data})
            results.append(row)
            if len(results) >= limit:
                break
        return results

    def _execute_optional_match(self, cypher: str, return_clause: str,
                                 where_clause: str, limit: int) -> List[Dict]:
        """Handle OPTIONAL MATCH … collect() pattern."""
        import re as _re
        primary_m = _re.search(
            r'^\s*MATCH\s+\(\s*(\w+)(?::(\w+))?\s*\)', cypher,
            _re.IGNORECASE | _re.DOTALL,
        )
        opt_m = _re.search(
            r'\bOPTIONAL\s+MATCH\s+\(\s*(\w+)(?::(\w+))?\s*\)\s*-\s*\[:(\w+)\]\s*->\s*\(\s*(\w+)(?::(\w+))?\s*\)',
            cypher, _re.IGNORECASE,
        )
        if not primary_m:
            return []
        anchor_var = primary_m.group(1)
        anchor_label = primary_m.group(2)
        return_items = self._split_return_items(return_clause)
        g = self._sandbox.graph._graph
        results: List[Dict] = []
        for nid, data in self._sandbox.graph.get_all_nodes():
            if anchor_label and data.get("_node_type") != anchor_label:
                continue
            anchor_data = dict(data)
            if where_clause and not self._eval_where(where_clause, {anchor_var: anchor_data}):
                continue
            # Collect optional targets
            opt_tgt_list: List[Dict] = []
            if opt_m:
                opt_rel_type = opt_m.group(3)
                opt_tgt_var = opt_m.group(4)
                opt_tgt_label = opt_m.group(5)
                for _, tgt_id, edge_data in g.out_edges(nid, data=True):
                    if edge_data.get("_rel_type") != opt_rel_type:
                        continue
                    tgt_data = dict(g.nodes.get(tgt_id, {}))
                    if opt_tgt_label and tgt_data.get("_node_type") != opt_tgt_label:
                        continue
                    opt_tgt_list.append(tgt_data)
            row: Dict = {}
            for expr, alias in return_items:
                collect_m = _re.match(r'collect\s*\((.+)\)$', expr.strip(),
                                      _re.IGNORECASE | _re.DOTALL)
                if collect_m and opt_m:
                    opt_tgt_var_name = opt_m.group(4)
                    collected = []
                    for tgt_data in opt_tgt_list:
                        val = self._eval_expr(
                            collect_m.group(1).strip(),
                            {anchor_var: anchor_data, opt_tgt_var_name: tgt_data},
                        )
                        if val is not None:
                            collected.append(val)
                    row[alias] = collected
                else:
                    row[alias] = self._eval_expr(expr, {anchor_var: anchor_data})
            results.append(row)
            if len(results) >= limit:
                break
        return results

    def _traverse_variable_length_rel(self, src_var: str, src_label: Optional[str],
                                       rel_type: str, min_depth: int, max_depth: int,
                                       tgt_var: str, tgt_label: Optional[str],
                                       return_items: List[tuple], where_clause: str,
                                       limit: int) -> List[Dict]:
        """BFS for variable-length paths (a)-[:REL*min..max]->(b)."""
        from collections import deque
        results: List[Dict] = []
        seen_pairs: set = set()
        g = self._sandbox.graph._graph
        for start_id, start_props in g.nodes(data=True):
            start_data = dict(start_props)
            if src_label and start_data.get("_node_type") != src_label:
                continue
            queue: deque = deque([(start_id, 1)])
            visited: set = {start_id}
            while queue:
                current_id, depth = queue.popleft()
                if depth > max_depth:
                    continue
                for _, next_id, edge_data in g.out_edges(current_id, data=True):
                    if edge_data.get("_rel_type") != rel_type:
                        continue
                    next_data = dict(g.nodes.get(next_id, {}))
                    if depth >= min_depth and next_id != start_id:
                        if (not tgt_label or next_data.get("_node_type") == tgt_label):
                            pair = (start_id, next_id)
                            if pair not in seen_pairs:
                                seen_pairs.add(pair)
                                var_map = {src_var: start_data, tgt_var: next_data}
                                if not where_clause or self._eval_where(where_clause, var_map):
                                    results.append(self._build_row(return_items, var_map))
                                    if len(results) >= limit:
                                        return results
                    if depth < max_depth and next_id not in visited:
                        visited.add(next_id)
                        queue.append((next_id, depth + 1))
        return results

    def _match_single_nodes(self, node_var: str, node_label: Optional[str],
                            return_items: List[tuple], where_clause: str,
                            limit: int) -> List[Dict]:
        """Scan graph nodes matching (node_var:node_label)."""
        results: List[Dict] = []
        for _nid, data in self._sandbox.graph.get_all_nodes():
            if node_label and data.get("_node_type") != node_label:
                continue
            var_map = {node_var: dict(data)}
            if where_clause and not self._eval_where(where_clause, var_map):
                continue
            results.append(self._build_row(return_items, var_map))
            if len(results) >= limit:
                break
        return results

    @staticmethod
    def _build_row(return_items: List[tuple], var_map: Dict) -> Dict:
        """Build a result row from [(expr, alias)] using _eval_expr for full expression support."""
        row: Dict = {}
        for expr, alias in return_items:
            row[alias] = HybridGraphService._eval_expr(expr, var_map)
        return row

    @staticmethod
    def _split_return_items(return_clause: str) -> List[tuple]:
        """Split RETURN clause respecting nested parentheses → [(expr, alias), …]."""
        import re as _re
        depth = 0
        current: List[str] = []
        items: List[str] = []
        for ch in return_clause:
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            if ch == ',' and depth == 0:
                items.append(''.join(current).strip())
                current = []
            else:
                current.append(ch)
        if current:
            items.append(''.join(current).strip())
        result: List[tuple] = []
        for item in items:
            item = item.strip()
            alias_m = _re.match(r'(.+?)\s+AS\s+(\w+)\s*$', item, _re.IGNORECASE)
            if alias_m:
                result.append((alias_m.group(1).strip(), alias_m.group(2).strip()))
            else:
                result.append((item, item))
        return result

    @staticmethod
    def _eval_expr(expr: str, var_map: Dict) -> Any:
        """Evaluate a single RETURN expression against a variable→data map.

        Handles:
          • var.prop
          • {key: var.prop, key2: var.prop2}
          • CASE WHEN var.prop IS NOT NULL THEN <expr> ELSE null END
          • coalesce(var.prop, var2.prop2, 'default')
        """
        import re as _re
        expr = expr.strip()

        # coalesce(arg1, arg2, ...) — return first non-None/non-empty value
        coalesce_m = _re.match(r'^coalesce\s*\((.+)\)\s*$', expr, _re.IGNORECASE | _re.DOTALL)
        if coalesce_m:
            # Split args on commas at depth 0
            depth = 0
            current: list = []
            args: list = []
            for ch in coalesce_m.group(1):
                if ch == '(':
                    depth += 1
                elif ch == ')':
                    depth -= 1
                if ch == ',' and depth == 0:
                    args.append(''.join(current).strip())
                    current = []
                else:
                    current.append(ch)
            if current:
                args.append(''.join(current).strip())
            for arg in args:
                arg = arg.strip()
                # String literal: 'value' or "value"
                lit_m = _re.match(r'^["\'](.*)["\']\ *$', arg)
                if lit_m:
                    return lit_m.group(1)  # return literal (may be empty string)
                val = HybridGraphService._eval_expr(arg, var_map)
                if val is not None and val != '':
                    return val
            # All args exhausted — return the last literal fallback or empty string
            for arg in reversed(args):
                lit_m = _re.match(r'^["\'](.*)["\']\ *$', arg.strip())
                if lit_m:
                    return lit_m.group(1)
            return ''

        # CASE WHEN var.prop IS NOT NULL THEN <then_expr> ELSE null END
        case_m = _re.match(
            r'CASE\s+WHEN\s+(\w+)\.(\w+)\s+IS\s+NOT\s+NULL\s+THEN\s+(.+?)\s+ELSE\s+null\s+END\s*$',
            expr, _re.IGNORECASE | _re.DOTALL,
        )
        if case_m:
            chk_var, chk_prop = case_m.group(1), case_m.group(2)
            then_expr = case_m.group(3).strip()
            if var_map.get(chk_var, {}).get(chk_prop) is not None:
                return HybridGraphService._eval_expr(then_expr, var_map)
            return None

        # Map literal: {key: var.prop, key2: var.prop2}
        map_m = _re.match(r'^\{(.+)\}$', expr, _re.DOTALL)
        if map_m:
            result: Dict = {}
            for pair in map_m.group(1).split(','):
                pair = pair.strip()
                km = _re.match(r'(\w+)\s*:\s*(\w+\.\w+|\w+)', pair)
                if km:
                    key, val_expr = km.group(1), km.group(2)
                    if '.' in val_expr:
                        v, p = val_expr.split('.', 1)
                        result[key] = var_map.get(v.strip(), {}).get(p.strip())
                    else:
                        result[key] = val_expr
            return result if result else None

        # var.prop
        dot_m = _re.match(r'^(\w+)\.(\w+)$', expr)
        if dot_m:
            return var_map.get(dot_m.group(1), {}).get(dot_m.group(2))

        # whole variable
        if _re.match(r'^\w+$', expr):
            return var_map.get(expr)
        return None

    @staticmethod
    def _extract_row(return_vars: List[str], var_map: Dict) -> Dict:
        """Build a result row from the RETURN variable list and node data map."""
        row: Dict = {}
        for rv in return_vars:
            rv = rv.strip()
            if '.' in rv:
                var, prop = rv.split('.', 1)
                row[rv] = var_map.get(var.strip(), {}).get(prop.strip())
            else:
                row[rv] = var_map.get(rv)
        return row

    @staticmethod
    def _eval_where(where_clause: str, var_map: Dict) -> bool:
        """Evaluate a simple WHERE clause against a variable→data map.

        Supports: =, IN [...], IS NOT NULL, IS NULL, AND.
        Unsupported predicates pass through (permissive — avoids false negatives).
        """
        import re as _re
        wc = where_clause.strip().lower()
        if wc == 'true':
            return True
        if wc == 'false':
            return False

        # Top-level OR: return True if any OR branch passes
        or_branches = _re.split(r'\bOR\b', where_clause, flags=_re.IGNORECASE)
        if len(or_branches) > 1:
            return any(
                HybridGraphService._eval_where(branch.strip(), var_map)
                for branch in or_branches
            )

        for clause in _re.split(r'\bAND\b', where_clause, flags=_re.IGNORECASE):
            clause = clause.strip()

            # var.prop IN [v1, v2, ...]
            in_m = _re.match(r'(\w+)\.(\w+)\s+IN\s+\[([^\]]*)\]', clause, _re.IGNORECASE)
            if in_m:
                var, prop, vals_str = in_m.group(1), in_m.group(2), in_m.group(3)
                values = [v.strip().strip('"\'') for v in vals_str.split(',') if v.strip()]
                actual = str(var_map.get(var, {}).get(prop, ""))
                if actual not in values:
                    return False
                continue

            # var.prop IS NOT NULL
            inn_m = _re.match(r'(\w+)\.(\w+)\s+IS\s+NOT\s+NULL', clause, _re.IGNORECASE)
            if inn_m:
                if var_map.get(inn_m.group(1), {}).get(inn_m.group(2)) is None:
                    return False
                continue

            # var.prop IS NULL
            inull_m = _re.match(r'(\w+)\.(\w+)\s+IS\s+NULL', clause, _re.IGNORECASE)
            if inull_m:
                if var_map.get(inull_m.group(1), {}).get(inull_m.group(2)) is not None:
                    return False
                continue

            # var.prop = "value" / 'value' / number
            eq_m = _re.match(
                r'(\w+)\.(\w+)\s*(?:=|=~)\s*["\']?([^"\']+?)["\']?\s*$', clause
            )
            if eq_m:
                var, prop, expected = eq_m.group(1), eq_m.group(2), eq_m.group(3).strip()
                actual = str(var_map.get(var, {}).get(prop, ""))
                if actual.upper() != expected.upper():
                    return False
                continue
            # Unsupported predicate — pass through
        return True

    def _canonical_id_from_record(self, record: Dict) -> Optional[str]:
        """Derive canonical node ID from a Neo4j result record."""
        # Try common key patterns from prod results
        node_type = record.get("label") or record.get("_node_type") or ""
        name = (record.get("name") or record.get("function_name")
                or record.get("param_name") or record.get("requirement_id")
                or record.get("test_case_id") or "")
        module = record.get("module") or "unknown"
        if node_type and name:
            return f"{node_type}:{name}:{module}"
        return None

    @classmethod
    def classify_tool(cls, tool_name: str) -> str:
        """Return 'sandbox', 'hybrid', or 'passthrough' for a given tool."""
        if tool_name in cls.SHALLOW_TOOLS:
            return "sandbox"
        elif tool_name in cls.DEEP_TOOLS:
            return "hybrid"
        return "sandbox"  # Default unclassified tools to sandbox


# ═════════════════════════════════════════════════════════════════════════
#  HybridTraversal — Seamless shadow → prod continuation
# ═════════════════════════════════════════════════════════════════════════

class HybridTraversal:
    """Traverse the enriched shadow graph with seamless prod continuation.

    When traversal hits a boundary leaf (a node with no further edges in
    NetworkX), it fires a continuation query against prod Neo4j from that
    node and stitches the results together. Re-entry detection prevents
    following stale prod paths for nodes that have sandbox overrides.

    Module-agnostic: works for any MCAL module or cross-module traversal.
    """

    def __init__(self, sandbox: "EphemeralSandbox", neo4j_driver=None,
                 workspace_id: str = "mcal"):
        self._sandbox = sandbox
        self._driver = neo4j_driver
        self._workspace_id = workspace_id
        # Shadow mask: canonical IDs that have sandbox overrides
        self._shadow_mask: set = set()
        self._build_shadow_mask()

    def _build_shadow_mask(self):
        """Collect canonical IDs of all sandbox-origin nodes."""
        self._shadow_mask = {
            nid for nid, data in self._sandbox.graph.get_all_nodes()
            if data.get("_origin") == "sandbox"
        }

    def _resolve_database(self) -> str:
        """Resolve Neo4j database name from workspace config."""
        try:
            from src.HybridRAG.code.neo4j_manager import get_instance_config
            cfg = get_instance_config(self._workspace_id)
            return cfg.database or "neo4j"
        except Exception:
            return "neo4j"

    def _is_boundary_leaf(self, node_id: str, direction: str,
                          rel_types: Optional[set] = None) -> bool:
        """Check if a node is a boundary leaf (no further edges in NetworkX).

        A node is a leaf if it exists in the graph but has no outgoing (or
        incoming, depending on direction) edges matching the filter criteria.
        """
        g = self._sandbox.graph._graph
        if node_id not in g:
            return False
        if direction in ("out", "both"):
            out_edges = [
                (s, t, d) for s, t, d in g.out_edges(node_id, data=True)
                if rel_types is None or d.get("_rel_type") in rel_types
            ]
            if not out_edges:
                return True
        if direction in ("in", "both"):
            in_edges = [
                (s, t, d) for s, t, d in g.in_edges(node_id, data=True)
                if rel_types is None or d.get("_rel_type") in rel_types
            ]
            if not in_edges:
                return True
        return False

    def traverse(
        self,
        start_id: str,
        direction: str = "out",
        max_depth: int = 4,
        rel_types: Optional[List[str]] = None,
        limit: int = 50,
    ) -> Dict[str, Any]:
        """Multi-hop traversal with seamless shadow → prod handoff.

        Parameters
        ----------
        start_id : str
            Canonical node ID to start traversal from.
        direction : str
            "out", "in", or "both".
        max_depth : int
            Maximum number of hops (across both shadow and prod combined).
        rel_types : list of str, optional
            Filter by relationship types. None = all types.
        limit : int
            Maximum total nodes to return.

        Returns
        -------
        dict with:
            "paths": list of path dicts (each: {nodes: [...], edges: [...]})
            "nodes": list of unique node dicts
            "boundary_continuations": int (how many times we jumped to prod)
            "truncated": bool
        """
        rel_set = set(rel_types) if rel_types else None
        visited: set = set()
        all_nodes: Dict[str, Dict] = {}
        all_edges: List[Dict] = []
        boundary_hits: List[str] = []  # leaf IDs that triggered prod continuation
        queue: List[tuple] = [(start_id, 0)]  # (node_id, current_depth)

        # Phase 1: BFS in NetworkX
        while queue and len(all_nodes) < limit:
            current_id, depth = queue.pop(0)
            if current_id in visited:
                continue
            visited.add(current_id)

            # Get node data from sandbox graph
            node_data = self._sandbox.graph.get_node(current_id)
            if node_data:
                all_nodes[current_id] = {
                    "node_id": current_id,
                    "node_type": node_data.get("_node_type", "Unknown"),
                    "origin": node_data.get("_origin", "unknown"),
                    "properties": {k: v for k, v in node_data.items()
                                   if not k.startswith("_")},
                }

            if depth >= max_depth:
                continue

            # Collect neighbors from NetworkX
            g = self._sandbox.graph._graph
            neighbors_found = False

            if direction in ("out", "both"):
                for _, tgt, edata in g.out_edges(current_id, data=True):
                    rt = edata.get("_rel_type", "RELATED_TO")
                    if rel_set and rt not in rel_set:
                        continue
                    neighbors_found = True
                    all_edges.append({
                        "source": current_id, "target": tgt,
                        "rel_type": rt, "origin": edata.get("_origin", "unknown"),
                    })
                    if tgt not in visited:
                        queue.append((tgt, depth + 1))

            if direction in ("in", "both"):
                for src, _, edata in g.in_edges(current_id, data=True):
                    rt = edata.get("_rel_type", "RELATED_TO")
                    if rel_set and rt not in rel_set:
                        continue
                    neighbors_found = True
                    all_edges.append({
                        "source": src, "target": current_id,
                        "rel_type": rt, "origin": edata.get("_origin", "unknown"),
                    })
                    if src not in visited:
                        queue.append((src, depth + 1))

            # Boundary leaf detection: node exists but has no qualifying edges
            if not neighbors_found and depth < max_depth:
                boundary_hits.append((current_id, depth))

        # Phase 2: Prod continuation from boundary leaves
        prod_results = {}
        if boundary_hits and self._driver:
            remaining_capacity = limit - len(all_nodes)
            if remaining_capacity > 0:
                prod_results = self._continue_in_prod(
                    boundary_hits, direction, max_depth,
                    rel_set, remaining_capacity, visited,
                )

        # Merge prod continuation results
        for nid, ndata in prod_results.get("nodes", {}).items():
            if nid not in all_nodes:
                all_nodes[nid] = ndata
        all_edges.extend(prod_results.get("edges", []))

        return {
            "start": start_id,
            "nodes": list(all_nodes.values()),
            "edges": all_edges,
            "total_nodes": len(all_nodes),
            "total_edges": len(all_edges),
            "boundary_continuations": len(prod_results.get("continued_from", [])),
            "continued_from": prod_results.get("continued_from", []),
            "truncated": len(all_nodes) >= limit,
        }

    def _continue_in_prod(
        self,
        boundary_leaves: List[tuple],
        direction: str,
        max_depth: int,
        rel_types: Optional[set],
        capacity: int,
        already_visited: set,
    ) -> Dict[str, Any]:
        """Fire continuation queries to prod Neo4j from boundary leaf nodes.

        For each boundary leaf, compute remaining depth and query prod.
        Re-entry check: if prod returns a node whose canonical ID is in the
        shadow mask, we stop following that path (sandbox has the truth).

        Parameters
        ----------
        boundary_leaves : list of (node_id, depth_at_leaf)
        direction, max_depth, rel_types : traversal params
        capacity : max nodes we can still accept
        already_visited : set of node IDs already in result

        Returns
        -------
        dict with "nodes", "edges", "continued_from"
        """
        db = self._resolve_database()
        nodes: Dict[str, Dict] = {}
        edges: List[Dict] = []
        continued_from: List[str] = []

        # Build direction clause
        if direction == "out":
            path_pattern = "(start)-[r*1..{depth}]->(n)"
        elif direction == "in":
            path_pattern = "(start)<-[r*1..{depth}]-(n)"
        else:
            path_pattern = "(start)-[r*1..{depth}]-(n)"

        for leaf_id, leaf_depth in boundary_leaves:
            if len(nodes) >= capacity:
                break

            remaining_depth = max_depth - leaf_depth
            if remaining_depth <= 0:
                continue

            # Extract name and module from canonical ID (e.g. "SRC_Function:Adc_Init:ADC")
            parts = leaf_id.split(":")
            if len(parts) < 3:
                continue
            node_type, name, module = parts[0], parts[1], parts[2]

            # Build relationship type filter
            rel_filter = ""
            if rel_types:
                rel_type_list = "|".join(rel_types)
                rel_filter = f":{rel_type_list}"

            # Build Cypher with literal depth (Neo4j 4.x limitation)
            safe_depth = int(remaining_depth)
            pattern = path_pattern.format(depth=safe_depth)
            # Replace [r*1..N] with relationship type filter if needed
            if rel_filter:
                pattern = pattern.replace("[r*1..","[r" + rel_filter + "*1..")

            cypher = f"""
            MATCH (start)
            WHERE (start.name = $name OR start.function_name = $name)
              AND toUpper(COALESCE(start.module, '')) = $module_upper
            WITH start LIMIT 1
            MATCH path = {pattern}
            WITH n, relationships(path) AS rels, length(path) AS hops
            ORDER BY hops
            LIMIT $node_limit
            RETURN
                toString(id(n)) AS neo4j_id,
                labels(n)[0] AS node_type,
                properties(n) AS props,
                [rel IN rels |
                    {{source: toString(id(startNode(rel))),
                      target: toString(id(endNode(rel))),
                      rel_type: type(rel)}}
                ] AS path_rels
            """

            try:
                with self._driver.session(database=db) as session:
                    result = session.run(
                        cypher,
                        name=name,
                        module_upper=module.upper(),
                        node_limit=int(capacity - len(nodes)),
                    )
                    continued_from.append(leaf_id)

                    for record in result:
                        # Build canonical ID for this prod node
                        props = record["props"] or {}
                        prod_canonical = EphemeralGraph._canonical_id(
                            record["node_type"], props
                        )

                        # Re-entry check: skip if this node is in shadow mask
                        if prod_canonical in self._shadow_mask:
                            logger.debug(
                                "[HybridTraversal] Re-entry blocked: %s (shadow has override)",
                                prod_canonical,
                            )
                            continue

                        # Skip already visited
                        if prod_canonical in already_visited:
                            continue

                        # Add node
                        if prod_canonical not in nodes:
                            nodes[prod_canonical] = {
                                "node_id": prod_canonical,
                                "node_type": record["node_type"],
                                "origin": "production_continuation",
                                "properties": {k: v for k, v in props.items()
                                               if v is not None},
                            }

                        # Add edges from path
                        for rel_info in (record["path_rels"] or []):
                            # Map Neo4j IDs to canonical IDs where possible
                            edge_entry = {
                                "source": rel_info["source"],
                                "target": rel_info["target"],
                                "rel_type": rel_info["rel_type"],
                                "origin": "production_continuation",
                                "_neo4j_source": rel_info["source"],
                                "_neo4j_target": rel_info["target"],
                            }
                            edges.append(edge_entry)

            except Exception as e:
                logger.warning(
                    "[HybridTraversal] Prod continuation failed from %s: %s",
                    leaf_id, e,
                )
                continue

        return {"nodes": nodes, "edges": edges, "continued_from": continued_from}

    def shortest_path(
        self,
        start_id: str,
        end_id: str,
        max_depth: int = 8,
        rel_types: Optional[List[str]] = None,
    ) -> Optional[Dict[str, Any]]:
        """Find shortest path between two nodes, spanning shadow + prod.

        Strategy:
        1. Try NetworkX first (if both nodes are in shadow graph).
        2. If end_id not in shadow, traverse from start to boundary leaves,
           then fire shortest_path from each leaf to end_id in prod.
        3. Stitch: shadow_path + prod_path, pick shortest total.

        Returns
        -------
        dict with "path_nodes", "path_edges", "total_hops", "segments"
        or None if no path found.
        """
        import networkx as nx
        g = self._sandbox.graph._graph
        rel_set = set(rel_types) if rel_types else None

        # Case 1: Both nodes in shadow graph — try direct NetworkX path
        if start_id in g and end_id in g:
            try:
                if rel_set:
                    # Filter graph to only edges with matching rel_types
                    def edge_filter(u, v):
                        return g.edges[u, v].get("_rel_type") in rel_set
                    view = nx.subgraph_view(g, filter_edge=edge_filter)
                    path = nx.shortest_path(view, start_id, end_id)
                else:
                    path = nx.shortest_path(g, start_id, end_id)
                return self._format_path(path, g, "shadow_only")
            except nx.NetworkXNoPath:
                pass  # Fall through to hybrid approach
            except nx.NodeNotFound:
                pass

        # Case 2: Hybrid — find boundary leaves reachable from start,
        #          then shortest_path in prod from leaf to end
        if not self._driver:
            return None

        # BFS from start to find boundary leaves (up to max_depth / 2)
        shadow_depth = min(max_depth // 2, 4)
        visited = set()
        queue = [(start_id, [start_id])]  # (node_id, path_so_far)
        leaf_paths: List[tuple] = []  # (leaf_id, path_to_leaf)

        while queue:
            current, path_so_far = queue.pop(0)
            if current in visited:
                continue
            visited.add(current)

            if len(path_so_far) > shadow_depth + 1:
                continue

            has_neighbors = False
            for _, tgt, edata in g.out_edges(current, data=True):
                if rel_set and edata.get("_rel_type") not in rel_set:
                    continue
                has_neighbors = True
                if tgt not in visited:
                    queue.append((tgt, path_so_far + [tgt]))

            if not has_neighbors and current != start_id:
                leaf_paths.append((current, path_so_far))

        if not leaf_paths:
            return None

        # Try prod shortest_path from each leaf to end_id
        db = self._resolve_database()
        best_result = None
        best_total_hops = float("inf")

        # Parse end_id to get name/module for matching
        end_parts = end_id.split(":")
        end_name = end_parts[1] if len(end_parts) >= 2 else end_id
        end_module = end_parts[2] if len(end_parts) >= 3 else ""

        rel_filter = ""
        if rel_types:
            rel_filter = ":" + "|".join(rel_types)

        for leaf_id, shadow_path in leaf_paths:
            leaf_parts = leaf_id.split(":")
            if len(leaf_parts) < 3:
                continue
            leaf_name, leaf_module = leaf_parts[1], leaf_parts[2]

            safe_depth = int(max_depth - len(shadow_path) + 1)
            if safe_depth <= 0:
                continue

            cypher = f"""
            MATCH (start), (end)
            WHERE (start.name = $start_name OR start.function_name = $start_name)
              AND toUpper(COALESCE(start.module, '')) = $start_module
              AND (end.name = $end_name OR end.function_name = $end_name
                   OR end.requirement_id = $end_name OR end.test_case_id = $end_name)
            WITH start, end LIMIT 1
            MATCH path = shortestPath((start)-[*..{safe_depth}]-(end))
            RETURN [n IN nodes(path) |
                    {{neo4j_id: toString(id(n)),
                      node_type: labels(n)[0],
                      props: properties(n)}}
                   ] AS path_nodes,
                   [r IN relationships(path) |
                    {{source: toString(id(startNode(r))),
                      target: toString(id(endNode(r))),
                      rel_type: type(r)}}
                   ] AS path_rels,
                   length(path) AS hops
            """

            try:
                with self._driver.session(database=db) as session:
                    result = session.run(
                        cypher,
                        start_name=leaf_name,
                        start_module=leaf_module.upper(),
                        end_name=end_name,
                    )
                    record = result.single()
                    if record:
                        total_hops = len(shadow_path) - 1 + record["hops"]
                        if total_hops < best_total_hops:
                            # Check re-entry for all nodes in prod path
                            prod_nodes = record["path_nodes"] or []
                            has_reentry = any(
                                EphemeralGraph._canonical_id(pn["node_type"], pn["props"])
                                in self._shadow_mask
                                for pn in prod_nodes[1:]  # skip start (it's the leaf)
                            )
                            if not has_reentry:
                                best_total_hops = total_hops
                                best_result = {
                                    "shadow_path": shadow_path,
                                    "prod_path_nodes": prod_nodes,
                                    "prod_path_rels": record["path_rels"],
                                    "leaf_id": leaf_id,
                                }
            except Exception as e:
                logger.debug(
                    "[HybridTraversal] shortest_path continuation from %s failed: %s",
                    leaf_id, e,
                )
                continue

        if not best_result:
            return None

        # Stitch shadow path + prod path
        path_nodes = []
        path_edges = []

        # Shadow segment
        for i, nid in enumerate(best_result["shadow_path"]):
            node_data = self._sandbox.graph.get_node(nid) or {}
            path_nodes.append({
                "node_id": nid,
                "node_type": node_data.get("_node_type", "Unknown"),
                "origin": node_data.get("_origin", "sandbox"),
                "properties": {k: v for k, v in node_data.items()
                               if not k.startswith("_")},
            })
            if i > 0:
                prev = best_result["shadow_path"][i - 1]
                edata = g.get_edge_data(prev, nid) or g.get_edge_data(nid, prev) or {}
                path_edges.append({
                    "source": prev, "target": nid,
                    "rel_type": edata.get("_rel_type", "RELATED_TO"),
                    "origin": "shadow",
                })

        # Prod segment (skip first node — it's the leaf, already in shadow path)
        for pn in best_result["prod_path_nodes"][1:]:
            canonical = EphemeralGraph._canonical_id(pn["node_type"], pn["props"])
            path_nodes.append({
                "node_id": canonical,
                "node_type": pn["node_type"],
                "origin": "production_continuation",
                "properties": {k: v for k, v in pn["props"].items()
                               if v is not None},
            })
        for pr in best_result["prod_path_rels"]:
            path_edges.append({
                "source": pr["source"],
                "target": pr["target"],
                "rel_type": pr["rel_type"],
                "origin": "production_continuation",
            })

        return {
            "path_nodes": path_nodes,
            "path_edges": path_edges,
            "total_hops": best_total_hops,
            "segments": [
                {"type": "shadow", "hops": len(best_result["shadow_path"]) - 1},
                {"type": "production", "hops": len(best_result["prod_path_nodes"]) - 1},
            ],
        }

    def _format_path(self, path: List[str], g, segment_type: str) -> Dict[str, Any]:
        """Format a NetworkX path into standard output."""
        path_nodes = []
        path_edges = []
        for i, nid in enumerate(path):
            data = dict(g.nodes.get(nid, {}))
            path_nodes.append({
                "node_id": nid,
                "node_type": data.get("_node_type", "Unknown"),
                "origin": data.get("_origin", "unknown"),
                "properties": {k: v for k, v in data.items() if not k.startswith("_")},
            })
            if i > 0:
                prev = path[i - 1]
                edata = g.get_edge_data(prev, nid) or {}
                path_edges.append({
                    "source": prev, "target": nid,
                    "rel_type": edata.get("_rel_type", "RELATED_TO"),
                    "origin": segment_type,
                })
        return {
            "path_nodes": path_nodes,
            "path_edges": path_edges,
            "total_hops": len(path) - 1,
            "segments": [{"type": segment_type, "hops": len(path) - 1}],
        }


# ═════════════════════════════════════════════════════════════════════════
#  SandboxQuerier — Unified query interface
# ═════════════════════════════════════════════════════════════════════════

class SandboxQuerier:
    """Query ephemeral graph + vectors, return unified results."""

    def __init__(self, sandbox: EphemeralSandbox, boost: float = 0.05):
        self._sandbox = sandbox
        self._boost = boost

    def search(self, query: str, top_k: int = 15, alpha: float = _DEFAULT_SEARCH_ALPHA) -> List[SearchResult]:
        """Combined graph keyword + vector semantic search."""
        graph_results = []
        vector_results = []

        if alpha > 0:
            keywords = re.split(r'[\s_,]+', query)
            keywords = [k for k in keywords if len(k) > 2]
            graph_results = self._sandbox.graph.keyword_search(keywords, top_k=top_k)
            for r in graph_results:
                r.score = r.score * alpha + self._boost

        if alpha < 1.0:
            vector_results = self._sandbox.vectors.search(query, top_k=top_k)
            for r in vector_results:
                r.score = r.score * (1 - alpha) + self._boost

        # Merge by node_id, keep highest score
        seen = {}
        for r in graph_results + vector_results:
            if r.node_id not in seen or r.score > seen[r.node_id].score:
                seen[r.node_id] = r
        merged = sorted(seen.values(), key=lambda x: -x.score)
        return merged[:top_k]

    def get_traceability(self, node_ids: List[str]) -> List[Dict]:
        return self._sandbox.graph.get_traceability(node_ids)
