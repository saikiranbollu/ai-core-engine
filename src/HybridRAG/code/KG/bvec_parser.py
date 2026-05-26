"""
BVEC (Boundary Value & Equivalence Class) Analysis Parser
==========================================================

Parses BVEC Analysis Report Excel files from the MCAL validation repository.

File naming convention:
    TC4xx_SW_MCAL_<Module>_BVEC_Analysis_Report.xlsx

Sheet types handled:
    - BVEC-InputParameter_<device>   → BVEC_InputParameter nodes
    - BVEC-OutputParameter           → BVEC_OutputParameter nodes
    - BVEC-ConfigParameter_<device>  → BVEC_ConfigParameter nodes

Each row with a non-null "Boundary Value" column becomes one node.
Merged/inherited cells (API name, parameter, eq class) are forward-filled
from previous rows.

Usage::

    from bvec_parser import parse_bvec_workbook

    entries = parse_bvec_workbook(
        xlsx_path="path/to/TC4xx_SW_MCAL_Eth_17_Leth_BVEC_Analysis_Report.xlsx",
        module="ETH_17_LETH",
    )
    # entries == {
    #     "BVEC_InputParameter": [...],
    #     "BVEC_OutputParameter": [...],
    #     "BVEC_ConfigParameter": [...],
    # }
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger("bvec_parser")

# ---------------------------------------------------------------------------
# Device extraction from sheet name
# ---------------------------------------------------------------------------
_DEVICE_RE = re.compile(r"_(TC\w+)$", re.IGNORECASE)


def _extract_device(sheet_name: str) -> Optional[str]:
    """Extract device variant from sheet name like 'BVEC-InputParameter_TC4DX'."""
    m = _DEVICE_RE.search(sheet_name)
    return m.group(1).upper() if m else None


def _clean(val: Any) -> Optional[str]:
    """Convert cell value to cleaned string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# Track UIDs to append suffix on collision
_uid_counter: Dict[str, int] = {}


def _make_uid(module: str, device: str, param_type: str, seq: int,
              eq_class: str, boundary_value: str) -> str:
    """
    Generate a deterministic UID for a BVEC entry.

    Format: {module}_{device}_{type}_{seq}_{eq_num}_{bv_code}[_N]
    Appends _N suffix if the base UID has already been used (handles
    multiple entries with same seq + eq_class + boundary_value type).
    """
    # Extract eq class number: "Eq Class - 2" → "2"
    eq_num = re.sub(r"[^0-9]", "", eq_class) if eq_class else "0"

    # Shorten boundary value: "Upper Limit (Max)" → "max", "Nominal" → "nom", "Lower Limit (Min)" → "min"
    bv_lower = (boundary_value or "").lower()
    if "nominal" in bv_lower:
        bv_code = "nom"
    elif "upper" in bv_lower or "max" in bv_lower:
        bv_code = "max"
    elif "lower" in bv_lower or "min" in bv_lower:
        bv_code = "min"
    else:
        bv_code = re.sub(r"[^a-z0-9]", "", bv_lower)[:8]

    base_uid = f"{module}_{device}_{param_type}_{seq}_eq{eq_num}_{bv_code}"

    # Handle collisions by appending a counter
    if base_uid in _uid_counter:
        _uid_counter[base_uid] += 1
        return f"{base_uid}_{_uid_counter[base_uid]}"
    else:
        _uid_counter[base_uid] = 0
        return base_uid


def _reset_uid_counter():
    """Reset the UID counter (call before each workbook parse)."""
    _uid_counter.clear()


# ---------------------------------------------------------------------------
# Input / Output Parameter Sheet Parser
# ---------------------------------------------------------------------------

def _parse_io_parameter_sheet(
    wb,
    sheet_name: str,
    module: str,
    device: str,
    param_type: str,  # "input" or "output"
) -> List[dict]:
    """
    Parse an InputParameter or OutputParameter sheet.

    InputParameter columns (0-indexed from col B=1):
        1: #, 2: API Name, 3: I/P parameter, 4: Range,
        5: Equivalence Class, 6: Class Name, 7: Class Range,
        8: Boundary Value, 9: Actual Values, 10: Test Case ID, 11: Remarks

    OutputParameter columns (0-indexed from col B=1):
        1: #, 2: API Name, 3: Output parameter name, 4: Range,
        5: Equivalence Class, 6: Class Range,
        7: Boundary Value, 8: Actual Values, 9: Test Case ID, 10: Remarks
        (No 'Class Name' column — shifted by 1)
    """
    if sheet_name not in wb.sheetnames:
        logger.debug("Sheet '%s' not found — skipping", sheet_name)
        return []

    ws = wb[sheet_name]
    nodes: List[dict] = []

    # Detect whether this is an Output sheet (different column layout)
    is_output = (param_type == "output")

    # State for forward-filling merged cells
    current_seq: int = 0
    current_api: Optional[str] = None
    current_param: Optional[str] = None
    current_range: Optional[str] = None
    current_eq_class: Optional[str] = None
    current_class_name: Optional[str] = None
    current_class_range: Optional[str] = None

    # Find header row (contains "Boundary Value")
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        vals = list(row)
        # Check across cols 7 and 8 for "Boundary Value"
        for col_idx in (7, 8):
            if len(vals) > col_idx and vals[col_idx] is not None and "Boundary Value" in str(vals[col_idx]):
                header_row = i
                break
        if header_row:
            break

    if header_row is None:
        header_row = 9 if is_output else 9
        logger.warning("Could not find header row in '%s', defaulting to row %d", sheet_name, header_row)

    data_start = header_row + 1

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        vals = list(row)
        while len(vals) < 12:
            vals.append(None)

        if is_output:
            # Output layout: 1:#, 2:API, 3:OutputParam, 4:Range,
            #   5:EqClass, 6:ClassRange, 7:BoundaryValue, 8:ActualValues, 9:TestCaseID, 10:Remarks
            seq_val = vals[1]
            api = _clean(vals[2])
            param = _clean(vals[3])
            range_val = _clean(vals[4])
            eq_class = _clean(vals[5])
            class_name = None  # Output sheet has no Class Name column
            class_range = _clean(vals[6])
            boundary_value = _clean(vals[7])
            actual_value = _clean(vals[8])
            test_case_id = _clean(vals[9])
            remarks = _clean(vals[10])
        else:
            # Input layout: 1:#, 2:API, 3:Param, 4:Range,
            #   5:EqClass, 6:ClassName, 7:ClassRange, 8:BV, 9:Actual, 10:TC, 11:Remarks
            seq_val = vals[1]
            api = _clean(vals[2])
            param = _clean(vals[3])
            range_val = _clean(vals[4])
            eq_class = _clean(vals[5])
            class_name = _clean(vals[6])
            class_range = _clean(vals[7])
            boundary_value = _clean(vals[8])
            actual_value = _clean(vals[9])
            test_case_id = _clean(vals[10])
            remarks = _clean(vals[11])

        # Update forward-fill state
        if seq_val is not None:
            try:
                current_seq = int(seq_val)
            except (ValueError, TypeError):
                pass
        if api is not None:
            current_api = api
            current_param = param
            current_range = range_val
        elif param is not None:
            current_param = param
            current_range = range_val
        if range_val is not None:
            current_range = range_val
        if eq_class is not None:
            current_eq_class = eq_class
            if class_name is not None:
                current_class_name = class_name
            if class_range is not None:
                current_class_range = class_range
        else:
            if class_name is not None:
                current_class_name = class_name
            if class_range is not None:
                current_class_range = class_range

        # Only create a node if there's a boundary value
        if boundary_value is None:
            continue

        # Skip if no API context yet (preamble rows)
        if current_api is None:
            continue

        uid = _make_uid(module, device, param_type, current_seq,
                        current_eq_class or "", boundary_value)

        node = {
            "uid": uid,
            "module": module,
            "device": device,
            "api_name": current_api,
            "parameter_name": current_param or "",
            "parameter_range": current_range or "",
            "equivalence_class": current_eq_class or "",
            "class_name": current_class_name or "",
            "class_range": current_class_range or "",
            "boundary_value": boundary_value,
            "actual_value": actual_value or "",
            "test_case_id": test_case_id or "",
            "remarks": remarks or "",
        }
        nodes.append(node)

    logger.info("Parsed %d entries from sheet '%s'", len(nodes), sheet_name)
    return nodes


# ---------------------------------------------------------------------------
# Config Parameter Sheet Parser
# ---------------------------------------------------------------------------

def _parse_config_parameter_sheet(
    wb,
    sheet_name: str,
    module: str,
    device: str,
) -> List[dict]:
    """
    Parse a ConfigParameter sheet.

    Header row columns (0-indexed from col B=1):
        1: #  (sequence number)
        2: Config parameter name
        3: Range
        4: Applicable API / sequence of APIs
        5: Equivalence Class
        6: Class Name
        7: Class Range
        8: Boundary Value
        9: Actual Values
        10: Test Case ID
        11: Remarks
    """
    if sheet_name not in wb.sheetnames:
        logger.debug("Sheet '%s' not found — skipping", sheet_name)
        return []

    ws = wb[sheet_name]
    nodes: List[dict] = []

    # State for forward-filling
    current_seq: int = 0
    current_param: Optional[str] = None
    current_range: Optional[str] = None
    current_apis: Optional[str] = None
    current_eq_class: Optional[str] = None
    current_class_name: Optional[str] = None
    current_class_range: Optional[str] = None

    # Find header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        vals = list(row)
        if len(vals) > 8 and vals[8] is not None and "Boundary Value" in str(vals[8]):
            header_row = i
            break

    if header_row is None:
        header_row = 14
        logger.warning("Could not find header row in '%s', defaulting to row %d", sheet_name, header_row)

    data_start = header_row + 1

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        vals = list(row)
        while len(vals) < 12:
            vals.append(None)

        seq_val = vals[1]
        param = _clean(vals[2])
        range_val = _clean(vals[3])
        apis = _clean(vals[4])
        eq_class = _clean(vals[5])
        class_name = _clean(vals[6])
        class_range = _clean(vals[7])
        boundary_value = _clean(vals[8])
        actual_value = _clean(vals[9])
        test_case_id = _clean(vals[10])
        remarks = _clean(vals[11])

        # Update forward-fill state
        if seq_val is not None:
            try:
                current_seq = int(seq_val)
            except (ValueError, TypeError):
                pass
        if param is not None:
            current_param = param
            current_range = range_val
        if range_val is not None:
            current_range = range_val
        if apis is not None:
            current_apis = apis
        if eq_class is not None:
            current_eq_class = eq_class
            if class_name is not None:
                current_class_name = class_name
            if class_range is not None:
                current_class_range = class_range
        else:
            if class_name is not None:
                current_class_name = class_name
            if class_range is not None:
                current_class_range = class_range

        # Only create a node if there's a boundary value
        if boundary_value is None:
            continue

        if current_param is None:
            continue

        uid = _make_uid(module, device, "config", current_seq,
                        current_eq_class or "", boundary_value)

        node = {
            "uid": uid,
            "module": module,
            "device": device,
            "api_name": current_apis or "",
            "parameter_name": current_param or "",
            "parameter_range": current_range or "",
            "equivalence_class": current_eq_class or "",
            "class_name": current_class_name or "",
            "class_range": current_class_range or "",
            "boundary_value": boundary_value,
            "actual_value": actual_value or "",
            "test_case_id": test_case_id or "",
            "remarks": remarks or "",
        }
        nodes.append(node)

    logger.info("Parsed %d entries from sheet '%s'", len(nodes), sheet_name)
    return nodes


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_bvec_workbook(
    xlsx_path: str | Path,
    module: str,
) -> Dict[str, List[dict]]:
    """
    Parse a BVEC Analysis Report workbook and return all entries by type.

    Args:
        xlsx_path: Path to the BVEC Excel file.
        module: MCAL module name (e.g. "ETH_17_LETH").

    Returns:
        Dict with keys "BVEC_InputParameter", "BVEC_OutputParameter",
        "BVEC_ConfigParameter", each containing a list of node dicts.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"BVEC file not found: {xlsx_path}")

    # Reset UID counter for fresh parse
    _reset_uid_counter()

    logger.info("Opening BVEC workbook: %s", xlsx_path.name)
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)

    result: Dict[str, List[dict]] = {
        "BVEC_InputParameter": [],
        "BVEC_OutputParameter": [],
        "BVEC_ConfigParameter": [],
    }

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("BVEC-InputParameter"):
            device = _extract_device(sheet_name)
            if device is None:
                logger.warning("Cannot extract device from sheet '%s'", sheet_name)
                continue
            entries = _parse_io_parameter_sheet(wb, sheet_name, module, device, "input")
            result["BVEC_InputParameter"].extend(entries)

        elif sheet_name.startswith("BVEC-OutputParameter"):
            # Output sheet is shared across devices — use "ALL" as device
            entries = _parse_io_parameter_sheet(wb, sheet_name, module, "ALL", "output")
            result["BVEC_OutputParameter"].extend(entries)

        elif sheet_name.startswith("BVEC-ConfigParameter"):
            device = _extract_device(sheet_name)
            if device is None:
                logger.warning("Cannot extract device from sheet '%s'", sheet_name)
                continue
            entries = _parse_config_parameter_sheet(wb, sheet_name, module, device)
            result["BVEC_ConfigParameter"].extend(entries)

    wb.close()

    total = sum(len(v) for v in result.values())
    logger.info(
        "BVEC parsing complete: %d InputParameter, %d OutputParameter, %d ConfigParameter (total: %d)",
        len(result["BVEC_InputParameter"]),
        len(result["BVEC_OutputParameter"]),
        len(result["BVEC_ConfigParameter"]),
        total,
    )

    return result


# ---------------------------------------------------------------------------
# CLI for standalone testing
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    if len(sys.argv) < 3:
        print("Usage: python bvec_parser.py <xlsx_path> <module_name>")
        print("Example: python bvec_parser.py path/to/BVEC.xlsx ETH_17_LETH")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    module_name = sys.argv[2]

    entries = parse_bvec_workbook(xlsx_path, module_name)

    for node_type, nodes in entries.items():
        print(f"\n{'='*60}")
        print(f"{node_type}: {len(nodes)} entries")
        print(f"{'='*60}")
        if nodes:
            # Print first 3 entries as sample
            for n in nodes[:3]:
                print(f"  uid: {n['uid']}")
                print(f"    api: {n['api_name']}, param: {n['parameter_name']}")
                print(f"    eq: {n['equivalence_class']}, bv: {n['boundary_value']}")
                print(f"    actual: {n['actual_value']}, tc: {n['test_case_id']}")
                print()
