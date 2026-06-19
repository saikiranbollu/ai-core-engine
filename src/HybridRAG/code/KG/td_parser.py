#!/usr/bin/env python3
"""
Test Data (TD) Excel Parser
============================

Parses TC4xx_SW_MCAL_TD_<Module>.xlsx workbooks into structured dictionaries
for ingestion into the knowledge graph.

The TD workbook contains:
  - IO parameter values: Test case IO parameters with expected values per config
  - Configuration details: Config metadata (index, file name, devices, regression)
  - HW Connections: Pin/signal mappings per device
  - Miscellaneous: Interface mode lookup + regression config mapping

Output: dict with keys:
  - "test_parameters": list of TD_TestParameter dicts
  - "configurations": list of TD_Configuration dicts
  - "hw_connections": list of TD_HWConnection dicts
  - "interface_modes": list of TD_InterfaceMode dicts
  - "regression_mappings": list of TD_RegressionMapping dicts
  - "metadata": document-level metadata (version, module, etc.)
"""
from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any, Optional

import openpyxl

logger = logging.getLogger("td_parser")

# Config ID pattern: AS460_TC49X_C021_P01
_CONFIG_ID_PATTERN = re.compile(r"AS460_(\w+?)_C(\d+)_P(\d+)")


def parse_td_workbook(xlsx_path: Path | str, module: str) -> dict[str, Any]:
    """Parse a Test Data workbook into structured data.

    Args:
        xlsx_path: Path to the TD Excel file.
        module: MCAL module name (e.g. "ETH_17_LETH").

    Returns:
        Dictionary with all parsed data from all sheets.
    """
    xlsx_path = Path(xlsx_path)
    logger.info("Opening TD workbook: %s", xlsx_path.name)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    result = {
        "test_parameters": [],
        "configurations": [],
        "hw_connections": [],
        "interface_modes": [],
        "regression_mappings": [],
        "metadata": _parse_metadata(wb, module),
    }

    # Parse each sheet
    if "Configuration details" in wb.sheetnames:
        result["configurations"] = _parse_configurations(wb["Configuration details"], module)
        logger.info("Parsed %d configurations", len(result["configurations"]))

    if "IO parameter values" in wb.sheetnames:
        result["test_parameters"] = _parse_io_parameters(wb["IO parameter values"], module)
        logger.info("Parsed %d test parameters", len(result["test_parameters"]))

    if "HW Connections" in wb.sheetnames:
        result["hw_connections"] = _parse_hw_connections(wb["HW Connections"], module)
        logger.info("Parsed %d HW connections", len(result["hw_connections"]))

    if "Miscellaneous" in wb.sheetnames:
        ifaces, regressions = _parse_miscellaneous(wb["Miscellaneous"], module)
        result["interface_modes"] = ifaces
        result["regression_mappings"] = regressions
        logger.info("Parsed %d interface modes, %d regression mappings",
                    len(ifaces), len(regressions))

    wb.close()

    logger.info(
        "TD parsing complete: %d params, %d configs, %d hw_conns, %d iface_modes, %d reg_maps",
        len(result["test_parameters"]),
        len(result["configurations"]),
        len(result["hw_connections"]),
        len(result["interface_modes"]),
        len(result["regression_mappings"]),
    )
    return result


def _parse_metadata(wb: openpyxl.Workbook, module: str) -> dict:
    """Extract document metadata from Cover page and Template version."""
    meta = {"module": module}

    if "Cover page" in wb.sheetnames:
        ws = wb["Cover page"]
        for r in range(1, min(20, ws.max_row + 1)):
            val_b = ws.cell(row=r, column=2).value
            val_c = ws.cell(row=r, column=3).value
            if val_b and "Maturity" in str(val_b):
                meta["maturity_status"] = str(val_c).strip() if val_c else None
            elif val_b and "Project Name" in str(val_b):
                meta["project_name"] = str(val_b).replace("Project Name:", "").strip()
            elif val_b and "Title" in str(val_b):
                meta["title"] = str(val_b).replace("Title:", "").strip()

        # Get latest version from revision history
        for r in range(6, min(50, ws.max_row + 1)):
            version = ws.cell(row=r, column=3).value
            if version and str(version).startswith("v"):
                meta["version"] = str(version).strip()
                date_val = ws.cell(row=r, column=2).value
                if date_val:
                    meta["version_date"] = str(date_val)[:10]
                break

    if "Template version" in wb.sheetnames:
        ws = wb["Template version"]
        for r in range(1, min(10, ws.max_row + 1)):
            val_a = ws.cell(row=r, column=1).value
            if val_a and "Template Version" in str(val_a):
                meta["template_version"] = str(ws.cell(row=r, column=4).value or "").strip()

    return meta


def _parse_configurations(ws, module: str) -> list[dict]:
    """Parse Configuration details sheet.

    Structure:
      Row 1: Headers (Index, Test Configuration File, DUT CORE Info, ASR Version,
              Device Information, ..., Regression Options, ..., Comments)
      Row 2: Sub-headers (AS460, TC499N_STD, TC4D9_COM, TC489_COM, TC457_RDR,
              SV, U, SV_U, Coverage, FI, Compilation, Manual)
      Row 3+: Data rows
    """
    configs = []

    # Read sub-headers from row 2 to identify device and regression columns
    # Device columns: 4-8 (AS460, TC499N_STD, TC4D9_COM, TC489_COM, TC457_RDR)
    # Regression columns: 9-15 (SV, U, SV_U, Coverage, FI, Compilation, Manual)
    device_cols = {}
    regression_cols = {}

    for col in range(3, ws.max_column + 1):
        header = ws.cell(row=2, column=col).value
        if not header:
            continue
        header = str(header).strip()
        # Identify device vs regression columns
        if header in ("AS460", "TC499N_STD", "TC4D9_COM", "TC489_COM", "TC457_RDR"):
            device_cols[col] = header
        elif header in ("SV", "U", "SV_U", "Coverage", "FI", "Compilation", "Manual"):
            regression_cols[col] = header

    # Parse data rows (start at row 3)
    for r in range(3, ws.max_row + 1):
        config_index = ws.cell(row=r, column=1).value
        config_file = ws.cell(row=r, column=2).value
        if not config_file:
            continue

        config_file = str(config_file).strip()

        # Extract config index from file name if not in col 1
        if not config_index:
            # Try to extract from name like "Eth_17_Leth_Config021"
            m = re.search(r"Config(\d+)", config_file)
            if m:
                config_index = m.group(1)

        # Collect device applicability
        devices = {}
        for col, dev_name in device_cols.items():
            val = ws.cell(row=r, column=col).value
            if val:
                devices[dev_name] = True

        # Collect regression options
        regression = {}
        for col, reg_name in regression_cols.items():
            val = ws.cell(row=r, column=col).value
            if val:
                regression[reg_name] = True

        # Comments (col 16 based on analysis)
        comments_col = max(regression_cols.keys()) + 1 if regression_cols else 16
        comments = ws.cell(row=r, column=comments_col).value
        comments_str = str(comments).strip() if comments else None

        configs.append({
            "config_index": str(config_index).zfill(3) if config_index else None,
            "config_file_name": config_file,
            "devices": devices,
            "regression_options": regression,
            "comments": comments_str,
            "module": module,
        })

    return configs


def _parse_io_parameters(ws, module: str) -> list[dict]:
    """Parse IO parameter values sheet.

    Structure:
      Row 2: Headers (Test Case ID, Input/Output Parameter, Data Type,
              Parameter Name, Condition, Param Set Values[P01], [P02], [P03])
      Each test case block:
        - First row (header): col 2 = TC ID, col 3 = None, cols 7+ = config IDs
        - Data rows: col 2 = TC ID, col 3 = direction, col 4 = type,
                     col 5 = param name, col 6 = condition, cols 7+ = values

    Returns list of dicts, one per IO parameter row with config_values map.
    """
    parameters = []
    current_tc = None
    current_configs = {}  # col_index -> config_id_string

    for r in range(3, ws.max_row + 1):
        tc_id = ws.cell(row=r, column=2).value
        if not tc_id:
            continue
        tc_id = str(tc_id).strip()

        direction = ws.cell(row=r, column=3).value

        if not direction:
            # This is a config header row for a new test case
            current_tc = tc_id
            current_configs = {}
            for col in range(7, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if v:
                    current_configs[col] = str(v).strip()
                # Stop scanning after many consecutive empty cols
                elif col > 7 and col not in current_configs:
                    # Check if next 5 cols are also empty
                    has_more = False
                    for check_col in range(col + 1, min(col + 6, ws.max_column + 1)):
                        if ws.cell(row=r, column=check_col).value:
                            has_more = True
                            break
                    if not has_more:
                        break
            continue

        # This is a data row
        direction = str(direction).strip()
        data_type = ws.cell(row=r, column=4).value
        param_name = ws.cell(row=r, column=5).value
        condition = ws.cell(row=r, column=6).value

        if not param_name:
            continue

        data_type = str(data_type).strip() if data_type else None
        param_name = str(param_name).strip()
        condition = str(condition).strip() if condition else None

        # Collect values for each config column
        config_values = {}
        for col, config_id in current_configs.items():
            val = ws.cell(row=r, column=col).value
            if val is not None:
                val_str = str(val).strip()
                # Split comma-separated config IDs that share a column
                # e.g. "AS460_TC4DX_C001_P01_BVEC,AS460_TC49X_C001_P01_BVEC"
                if ',' in config_id:
                    for single_id in config_id.split(','):
                        single_id = single_id.strip()
                        if single_id:
                            config_values[single_id] = val_str
                else:
                    config_values[config_id] = val_str

        parameters.append({
            "test_case_id": tc_id,
            "direction": direction,
            "data_type": data_type,
            "parameter_name": param_name,
            "condition": condition,
            "config_values": config_values,
            "module": module,
        })

    return parameters


def _parse_hw_connections(ws, module: str) -> list[dict]:
    """Parse HW Connections sheet.

    Structure:
      Row 2: Device section headers (e.g. "TC4DX Reduced Media Independent...")
      Row 3: Sub-headers (pin number, characteristics)
      Row 4: Characteristic field names (PinDirection, PinInitialMode, etc.)
      Row 5+: Data rows (signal name, pin, characteristics per device)

    Each device section has columns: signal_name, pin_number, then 6 characteristics.
    """
    connections = []

    # Identify device sections from row 2
    device_sections = []  # list of (start_col, device_name)
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=col).value
        if v and ("Interface" in str(v) or "Signal" in str(v) or "signal" in str(v)):
            # Extract device name from header
            header = str(v).strip()
            # e.g. "TC4DX Reduced Media Independent Interface (RMII) Signals (LETH0_PORT0)"
            device_match = re.match(r"\s*(TC\w+)", header)
            device = device_match.group(1) if device_match else f"Device_col{col}"
            device_sections.append((col, device, header))

    if not device_sections:
        # Fallback: try to detect from row 3 "pin number" occurrences
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=3, column=col).value
            if v and "pin number" in str(v).lower():
                device_sections.append((col, f"Device_{col}", ""))

    # Get characteristic field names from row 4
    char_fields = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=col).value
        if v:
            char_fields.append((col, str(v).strip()))

    # Parse data rows for each device section
    for section_idx, (start_col, device, section_header) in enumerate(device_sections):
        # Signal name col = start_col, pin col = start_col + 1
        # Characteristics start at start_col + 2
        signal_col = start_col
        pin_col = start_col + 1

        # Find characteristic columns for this section
        # They follow the pin column
        section_chars = []
        for col, field_name in char_fields:
            if section_idx < len(device_sections) - 1:
                next_start = device_sections[section_idx + 1][0]
                if start_col < col < next_start:
                    section_chars.append((col, field_name))
            else:
                if col > start_col:
                    section_chars.append((col, field_name))

        # Parse data rows (row 5 onwards)
        for r in range(5, ws.max_row + 1):
            signal = ws.cell(row=r, column=signal_col).value
            if not signal:
                continue
            signal = str(signal).strip()

            # Skip section header/separator rows (not actual signal data).
            # Real signals: "LETH0_P0_RMIIB_TXD0", "ETH_TXD0"
            # Headers: "pin number", "Interface", "LETH0_PORT0", "LETH0_PORT1"
            signal_lower = signal.lower()
            if (signal_lower == "pin number"
                    or signal_lower == "interface"
                    or re.match(r"^leth\d+_port\d+$", signal_lower)):
                continue

            pin = ws.cell(row=r, column=pin_col).value
            pin = str(pin).strip() if pin else None

            characteristics = {}
            for col, field_name in section_chars:
                val = ws.cell(row=r, column=col).value
                if val:
                    characteristics[field_name] = str(val).strip()

            connections.append({
                "signal_name": signal,
                "pin_number": pin,
                "device": device,
                "section_header": section_header,
                "characteristics": characteristics,
                "module": module,
            })

    return connections


def _parse_miscellaneous(ws, module: str) -> tuple[list[dict], list[dict]]:
    """Parse Miscellaneous sheet.

    Contains:
      1. Interface mode lookup table (rows 2-10ish)
      2. Regression configuration mappings per device (rows 13+)

    Returns:
        Tuple of (interface_modes, regression_mappings)
    """
    interface_modes = []
    regression_mappings = []

    # Section 1: Interface modes (row 2 = header, row 3+ = data until blank)
    header_row = 2
    headers = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v:
            headers.append((col, str(v).strip()))

    for r in range(3, ws.max_row + 1):
        interface = ws.cell(row=r, column=1).value
        if not interface:
            break  # End of interface mode section

        mode = {"interface": str(interface).strip(), "module": module}
        for col, hdr in headers[1:]:  # Skip first (Interface) header
            val = ws.cell(row=r, column=col).value
            if val is not None:
                # Normalize header to key
                key = hdr.lower().replace(" ", "_").replace("(", "").replace(")", "")
                mode[key] = str(val).strip()
        interface_modes.append(mode)

    # Section 2: Regression mappings (detect by looking for "Configuration" in cells)
    # These are device-specific regression config tables
    current_device = None
    in_data = False

    for r in range(10, ws.max_row + 1):
        # Check for device header rows
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=col).value
            if v and "Regression Configuration" in str(v):
                # Extract device name
                dev_match = re.match(r"(TC\w+)", str(v).strip())
                current_device = dev_match.group(1) if dev_match else str(v).strip()
                in_data = False
                break
            elif v and "Test combination" in str(v):
                in_data = True
                break
        else:
            # Data row
            if in_data and current_device:
                test_combo = ws.cell(row=r, column=1).value
                config_nums = ws.cell(row=r, column=2).value
                if test_combo and config_nums:
                    regression_mappings.append({
                        "device": current_device,
                        "test_combination": str(test_combo).strip(),
                        "config_numbers": str(config_nums).strip(),
                        "module": module,
                    })

                # Check right-side table (some rows have 2 device tables side by side)
                test_combo_r = ws.cell(row=r, column=3).value
                config_nums_r = ws.cell(row=r, column=4).value
                if test_combo_r and config_nums_r:
                    # Need to identify which device this belongs to
                    # Check row 2 headers for right-side device
                    right_device = None
                    for col in range(3, ws.max_column + 1):
                        rv = ws.cell(row=r - 1, column=col).value
                        if rv and "Regression" in str(rv):
                            dev_m = re.match(r"(TC\w+)", str(rv).strip())
                            right_device = dev_m.group(1) if dev_m else None
                            break

                    regression_mappings.append({
                        "device": right_device or "Unknown",
                        "test_combination": str(test_combo_r).strip(),
                        "config_numbers": str(config_nums_r).strip(),
                        "module": module,
                    })

    return interface_modes, regression_mappings


# ---------------------------------------------------------------------------
# CLI for standalone testing
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s - %(message)s")

    if len(sys.argv) < 2:
        print("Usage: python td_parser.py <xlsx_path> [module]")
        sys.exit(1)

    xlsx = Path(sys.argv[1])
    mod = sys.argv[2] if len(sys.argv) > 2 else "ETH_17_LETH"

    data = parse_td_workbook(xlsx, mod)

    print(f"\n{'='*60}")
    print(f"  TD Parse Summary — {mod}")
    print(f"{'='*60}")
    print(f"  Test parameters: {len(data['test_parameters'])}")
    print(f"  Configurations:  {len(data['configurations'])}")
    print(f"  HW Connections:  {len(data['hw_connections'])}")
    print(f"  Interface modes: {len(data['interface_modes'])}")
    print(f"  Regression maps: {len(data['regression_mappings'])}")
    print(f"  Metadata:        {data['metadata']}")
    print(f"{'='*60}")

    # Show some samples
    if data["test_parameters"]:
        print("\nSample test parameters (first 3):")
        for p in data["test_parameters"][:3]:
            cv = p["config_values"]
            print(f"  {p['test_case_id']} | {p['direction']} | {p['parameter_name']} | "
                  f"{p['condition']} | configs: {len(cv)} | vals: {list(cv.items())[:2]}")

    if data["configurations"]:
        print("\nSample configurations (first 3):")
        for c in data["configurations"][:3]:
            print(f"  {c['config_index']}: {c['config_file_name']} | "
                  f"devices: {list(c['devices'].keys())} | "
                  f"regression: {list(c['regression_options'].keys())}")

    if data["hw_connections"]:
        print("\nSample HW connections (first 3):")
        for h in data["hw_connections"][:3]:
            print(f"  {h['device']}: {h['signal_name']} @ {h['pin_number']} | "
                  f"chars: {list(h['characteristics'].keys())[:3]}")

    if data["interface_modes"]:
        print("\nInterface modes:")
        for m in data["interface_modes"]:
            print(f"  {m['interface']} @ {m.get('speedmbps', '?')} Mbps -> "
                  f"PhyConfig={m.get('ethtest_phyconfigdecimal', '?')}")
